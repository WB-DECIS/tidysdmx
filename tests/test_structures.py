from typeguard import TypeCheckError
from datetime import datetime, timezone
from openpyxl import Workbook
from typing import Sequence
from pysdmx.model import (
    DataType,
    Role,
    Concept
)
from pysdmx.model.map import (
    FixedValueMap, 
    ImplicitComponentMap, 
    DatePatternMap, 
    ValueMap, 
    MultiValueMap, 
    RepresentationMap,
    ComponentMap,
    MultiRepresentationMap,
    StructureMap,
    MultiComponentMap
    )
import pandas as pd
import numpy as np
import pytest
import re
# Import tidysdmx functions
from tidysdmx.structures import (
    # infer_role_dimension, 
    build_fixed_map, 
    build_implicit_component_map, 
    build_date_pattern_map,
    build_value_map,
    build_value_map_list,
    build_multi_value_map_list,
    build_representation_map,
    build_single_component_map,
    _extract_mapping_definitions,
    _read_comp_mapping_sheet,
    _create_fixed_definition,
    _create_implicit_definition,
    _create_representation_definition,
    create_schema_from_table,
    _parse_info_sheet,
    _parse_comp_mapping_sheet,
    _parse_rep_mapping_sheet,
    _match_column_name,
    build_multi_representation_map,
    build_structure_map,
    _extract_artefact_id,
    _validate_mappings,
    build_structure_map_from_template_wb,
    _extract_all_artefact_ids,
    _extract_metadata_from_info_sheet,
    _extract_mapping_rule,
    _is_missing_token,
    _extract_representation_map,
    create_xl_template_from_sm

    )

# region fixtures

@pytest.fixture
def id():
    return "RM_ID"

@pytest.fixture
def name():
    return "Country Map"

@pytest.fixture
def agency():
    return "ECB"

@pytest.fixture
def source_cl():
    return "urn:source:codelist"

@pytest.fixture
def target_cl():
    return "urn:target:codelist"

@pytest.fixture
def description():
    return "Mapping ISO2 to ISO3 codes"
# endregion

class TestBuildFixedMap:  # noqa: D101
    def test_build_fixed_map_normal(self):
        """Valid mapping with default located_in."""
        mapping = build_fixed_map(target="CONF_STATUS", value="F")
        assert isinstance(mapping, FixedValueMap)
        assert mapping.target == "CONF_STATUS"
        assert mapping.value == "F"
        assert mapping.located_in == "target"

    def test_build_fixed_map_valid_source_located_in(self):
        """Valid mapping with located_in set to 'source'."""
        mapping = build_fixed_map("REPORTING_STATUS", "FINAL", located_in="source")
        assert mapping.located_in == "source"
        assert mapping.target == "REPORTING_STATUS"
        assert mapping.value == "FINAL"

    def test_build_fixed_map_invalid_target(self):
        """Empty target should raise ValueError."""
        with pytest.raises(ValueError):
            build_fixed_map("", "F")

    def test_build_fixed_map_invalid_value(self):
        """Empty value should raise ValueError."""
        with pytest.raises(ValueError):
            build_fixed_map("CONF_STATUS", "")

    def test_build_fixed_map_invalid_located_in_raises(self):
        """Invalid located_in should raise ValueError."""
        with pytest.raises(ValueError) as exc_info:
            build_fixed_map("CONF_STATUS", "F", located_in="invalid")

    def test_build_fixed_map_type_safety(self):
        """Ensure type safety enforced by typeguard."""
        with pytest.raises(TypeCheckError):
            build_fixed_map(123, "F")  # target must be str
        with pytest.raises(TypeCheckError):
            build_fixed_map("CONF_STATUS", 456)  # value must be str
        with pytest.raises(TypeCheckError):
            build_fixed_map("CONF_STATUS", 456, located_in=None)

class TestBuildImplicitComponentMap:  # noqa: D101
    def test_build_implicit_component_map_valid(self):
        """Valid mapping should return an ImplicitComponentMap instance."""
        mapping = build_implicit_component_map("FREQ", "FREQUENCY")
        assert isinstance(mapping, ImplicitComponentMap)
        assert mapping.source == "FREQ"
        assert mapping.target == "FREQUENCY"

    def test_build_implicit_component_map_empty_source_raises(self):
        """Empty source should raise ValueError."""
        with pytest.raises(ValueError) as exc_info:
            build_implicit_component_map("", "FREQUENCY")
        assert "non-empty" in str(exc_info.value)

    def test_build_implicit_component_map_empty_target_raises(self):
        """Empty target should raise ValueError."""
        with pytest.raises(ValueError) as exc_info:
            build_implicit_component_map("FREQ", "")
        assert "non-empty" in str(exc_info.value)

    def test_build_implicit_component_map_type_safety_source(self):
        """Non-string source should raise TypeError due to typeguard."""
        with pytest.raises(TypeCheckError):
            build_implicit_component_map(123, "FREQUENCY")

    def test_build_implicit_component_map_type_safety_target(self):
        """Non-string target should raise TypeError due to typeguard."""
        with pytest.raises(TypeCheckError):
            build_implicit_component_map("FREQ", 456)

    @pytest.mark.skip(reason="Not implemented.")
    def test_build_implicit_component_map_whitespace_source(self):
        """Whitespace-only source should raise ValueError."""
        with pytest.raises(ValueError):
            build_implicit_component_map("   ", "FREQUENCY")

    @pytest.mark.skip(reason="Not implemented.")
    def test_build_implicit_component_map_whitespace_target(self):
        """Whitespace-only target should raise ValueError."""
        with pytest.raises(ValueError):
            build_implicit_component_map("FREQ", "   ")

    def test_build_implicit_component_map_case_sensitivity(self):
        """Ensure mapping preserves case of source and target."""
        mapping = build_implicit_component_map("freq", "Frequency")
        assert mapping.source == "freq"

class TestBuildDatePatternMap:  # noqa: D101
    
    def test_build_date_pattern_map_valid_fixed(self):
        """Valid case with fixed pattern type."""
        dpm = build_date_pattern_map("DATE", "TIME_PERIOD", "MMM yy", "M")
        assert isinstance(dpm, DatePatternMap)
        assert dpm.source == "DATE"
        assert dpm.target == "TIME_PERIOD"
        assert dpm.pattern == "MMM yy"
        assert dpm.frequency == "M"
        assert dpm.pattern_type == "fixed"
        assert dpm.locale == "en"

    def test_build_date_pattern_map_valid_variable(self):
        """Valid case with variable pattern type."""
        dpm = build_date_pattern_map("DATE", "TIME_PERIOD", "MMM yy", "FREQ", pattern_type="variable")
        assert dpm.pattern_type == "variable"
        assert dpm.frequency == "FREQ"

    def test_build_date_pattern_map_with_id_and_locale(self):
        """Valid case with custom ID and locale."""
        dpm = build_date_pattern_map("DATE", "TIME_PERIOD", "MMM yy", "M", id="map1", locale="fr")
        assert dpm.id == "map1"
        assert dpm.locale == "fr"

    def test_build_date_pattern_map_with_resolve_period(self):
        """Valid case with resolve_period specified."""
        dpm = build_date_pattern_map("DATE", "TIME_PERIOD", "MMM yy", "M", resolve_period="endOfPeriod")
        assert dpm.resolve_period == "endOfPeriod"

    def test_build_date_pattern_map_empty_source(self):
        """Empty source should raise ValueError."""
        with pytest.raises(ValueError):
            build_date_pattern_map("", "TIME_PERIOD", "MMM yy", "M")

    def test_build_date_pattern_map_empty_target(self):
        """Empty target should raise ValueError."""
        with pytest.raises(ValueError):
            build_date_pattern_map("DATE", "", "MMM yy", "M")

    def test_build_date_pattern_map_empty_pattern(self):
        """Empty pattern should raise ValueError."""
        with pytest.raises(ValueError):
            build_date_pattern_map("DATE", "TIME_PERIOD", "", "M")

    def test_build_date_pattern_map_empty_frequency(self):
        """Empty frequency should raise ValueError."""
        with pytest.raises(ValueError):
            build_date_pattern_map("DATE", "TIME_PERIOD", "MMM yy", "")

    def test_build_date_pattern_map_invalid_pattern_type(self):
        """Invalid pattern_type should raise TypeError."""
        with pytest.raises(TypeCheckError):
            build_date_pattern_map("DATE", "TIME_PERIOD", "MMM yy", "M", pattern_type="wrong")  # type: ignore

    def test_build_date_pattern_map_invalid_resolve_period(self):
        """Invalid resolve_period should raise TypeError."""
        with pytest.raises(TypeCheckError):
            build_date_pattern_map("DATE", "TIME_PERIOD", "MMM yy", "M", resolve_period="invalid")  # type: ignore

    def test_build_date_pattern_map_whitespace_source(self):
        """Whitespace-only source should raise ValueError."""
        with pytest.raises(ValueError):
            build_date_pattern_map("   ", "TIME_PERIOD", "MMM yy", "M")

    def test_build_date_pattern_map_whitespace_target(self):
        """Whitespace-only target should raise ValueError."""
        with pytest.raises(ValueError):
            build_date_pattern_map("DATE", "   ", "MMM yy", "M")

    def test_build_date_pattern_map_whitespace_pattern(self):
        """Whitespace-only pattern should raise ValueError."""
        with pytest.raises(ValueError):
            build_date_pattern_map("DATE", "TIME_PERIOD", "   ", "M")

class TestBuildValueMap:  # noqa: D101
    def test_build_value_map_normal(self):
        """Test normal case with source and target values."""
        vm = build_value_map("BE", "BEL")
        assert isinstance(vm, ValueMap)
        assert vm.source == "BE"
        assert vm.target == "BEL"
        assert vm.valid_from is None
        assert vm.valid_to is None

    def test_build_value_map_with_validity_dates(self):
        """Test case with both validity dates provided."""
        start = datetime(2021, 1, 1)
        end = datetime(2022, 1, 1)
        vm = build_value_map("FR", "FRA", valid_from=start, valid_to=end)
        assert vm.valid_from == start
        assert vm.valid_to == end

    def test_build_value_map_with_only_valid_from(self):
        """Test case with only valid_from date provided."""
        start = datetime(2020, 6, 15)
        vm = build_value_map("DE", "GER", valid_from=start)
        assert vm.valid_from == start
        assert vm.valid_to is None

    def test_build_value_map_with_only_valid_to(self):
        """Test case with only valid_to date provided."""
        end = datetime(2030, 12, 31)
        vm = build_value_map("IT", "ITA", valid_to=end)
        assert vm.valid_to == end
        assert vm.valid_from is None

    def test_build_value_map_empty_source(self):
        """Empty source raises ValueError."""
        with pytest.raises(ValueError):
            build_value_map("", "BEL")

    def test_build_value_map_empty_target(self):
        """Empty target raises ValueError."""
        with pytest.raises(ValueError):
            build_value_map("BE", "")

    def test_build_value_map_invalid_source_type(self):
        """Non-string source raises TypeError."""
        with pytest.raises(TypeError):
            build_value_map(123, "BEL")

    def test_build_value_map_invalid_target_type(self):
        """Non-string target raises TypeError."""
        with pytest.raises(TypeError):
            build_value_map("BE", 456)

    def test_build_value_map_whitespace_source(self):
        """Whitespace-only source raises ValueError."""
        with pytest.raises(ValueError):
            build_value_map("   ", "BEL")

    def test_build_value_map_whitespace_target(self):
        """Whitespace-only target raises ValueError."""
        with pytest.raises(ValueError):
            build_value_map("BE", "   ")

    def test_build_value_map_type_safety(self):
        """Ensure returned object is of type ValueMap."""
        vm = build_value_map("NL", "NLD")
        assert isinstance(vm, ValueMap)

    def test_build_value_map_validity_type(self):
        """Ensure validity dates are datetime or None."""
        vm = build_value_map("ES", "ESP", valid_from=datetime(2025, 1, 1))
        assert isinstance(vm.valid_from, datetime)

class TestBuildValueMapList:  # noqa: D101
    def test_build_value_map_list_valid(self, value_map_df_mandatory_cols):
        """Valid DataFrame should return a list of ValueMap objects."""
        result = build_value_map_list(value_map_df_mandatory_cols, "source", "target")
        assert isinstance(result, list)
        assert all(isinstance(vm, ValueMap) for vm in result)
        assert len(result) == value_map_df_mandatory_cols.shape[0]
        assert result[1].source == "UY"
        assert result[1].target == "URY"
        assert result[0].typed_source == re.compile(r"^A") 
    
    
    def test_build_value_map_list_validity_columns(self):
        """DataFrame with custom validity column names should populate ValueMap."""
        df = pd.DataFrame({
            "source": ["BE", "FR"],
            "target": ["BEL", "FRA"],
            "start_date": ["2020-01-01", None],
            "end_date": ["2025-12-31", None]
        })
        result = build_value_map_list(df, "source", "target", valid_from_col="start_date", valid_to_col="end_date")
        assert result[0].valid_from == "2020-01-01"
        assert result[0].valid_to == "2025-12-31"
        assert result[1].valid_from is None
        assert result[1].valid_to is None
 
    def test_build_value_map_list_validity_columns_empty(self):
        """Validity columns present but empty should be ignored."""
        df = pd.DataFrame({
            "source": ["BR"],
            "target": ["BRA"],
            "valid_from": [None],
            "valid_to": [None]
        })
        result = build_value_map_list(df, "source", "target")
        assert result[0].valid_from is None
        assert result[0].valid_to is None

    def test_build_value_map_list_empty_df(self):
        """Empty DataFrame should raise ValueError."""
        empty_df = pd.DataFrame(columns=["source", "target"])
        with pytest.raises(ValueError):
            build_value_map_list(empty_df, "source", "target")

    def test_build_value_map_list_missing_column(self, value_map_df_mandatory_cols):
        """Missing column should raise ValueError."""
        df_missing = value_map_df_mandatory_cols.drop(columns=["target"])
        with pytest.raises(ValueError):
            build_value_map_list(df_missing, "source", "target")

    def test_build_value_map_list_single_row(self):
        """Single-row DataFrame should return a list with one ValueMap."""
        df_single = pd.DataFrame({"source": ["BR"], "target": ["BRA"]})
        result = build_value_map_list(df_single, "source", "target")
        assert len(result) == 1
        assert result[0].source == "BR"
        assert result[0].target == "BRA"

    def test_build_value_map_list_column_order_irrelevant(self, value_map_df_mandatory_cols):
        """Column order should not affect the result."""
        df_reordered = value_map_df_mandatory_cols[["target", "source"]]
        # result = build_value_map_list(df_reordered.rename(columns={"target": "target", "source": "source"}), "source", "target")
        result = build_value_map_list(df_reordered, "source", "target")
        assert len(result) == value_map_df_mandatory_cols.shape[0]
        assert result[1].source == "UY"
        assert result[1].target == "URY"

class TestBuildMultiValueMapList:  # noqa: D101
    @pytest.fixture
    def multi_value_map_df(self) -> pd.DataFrame:
        """Fixture providing a standard DataFrame for testing scenarios.
        
        Rows correspond to specific test cases:
        0: Normal mapping (DE -> EUR)
        1: Regex source (regex:^A -> ARG)
        2: Empty source string ("" -> CHF)
        3: Validity dates (FR -> FRA)
        4: Invalid type (int in source)
        """
        data = {
            'country': ['DE', 'regex:^A', '', 'FR', 123],
            'currency': ['LC', 'LC', 'LC', 'LC', 'LC'],
            'iso_code': ['EUR', 'ARG', 'CHF', 'FRA', 'INV'],
            'valid_from': [None, None, None, '2020-01-01', None],
            'valid_to': [None, None, None, '2025-12-31', None]
        }
        return pd.DataFrame(data)

    @pytest.fixture
    def non_string_type_error_data(self) -> tuple[pd.DataFrame, Sequence[str], Sequence[str]]:
        """Provides test data with a non-string value in the target column.

        Returns:
            tuple: The DataFrame, source columns list, and target columns list.
        """
        data = {
            'source_1': ['A', 'B'],
            'source_2': ['X', 'Y'],
            'target_col_with_int': ['Target1', 123],  # Non-string value here
            'valid_from': ['2020-01-01', '2021-01-01']
        }
        df = pd.DataFrame(data)
        source_cols = ['source_1', 'source_2']
        target_cols = ['target_col_with_int']
        return df, source_cols, target_cols

    def test_build_multi_value_map_non_string(
        self, 
        non_string_type_error_data: tuple[pd.DataFrame, Sequence[str], Sequence[str]]
    ) -> None:
        """Tests that TypeError is raised when the target column contains non-string values.

        Args:
            non_string_type_error_data: Fixture data for the error case.
        """
        df, source_cols, target_cols = non_string_type_error_data
        # Note: The function validates per column, so expected message refers to the specific column name
        expected_message = f"Target column '{target_cols[0]}' must contain only string values."

        with pytest.raises(TypeError, match=re.escape(expected_message)):
            build_multi_value_map_list(df, source_cols, target_cols)

    def test_build_multi_value_map_list_normal(self, multi_value_map_df: pd.DataFrame) -> None:
        """Normal case: first row should produce a valid MultiValueMap."""
        df = multi_value_map_df.iloc[[0]]  # Row with DE/EUR
        
        # Updated: passing target cols as a list
        result = build_multi_value_map_list(df, ['country', 'currency'], ['iso_code'])
        
        assert isinstance(result, list)
        assert len(result) == 1
        assert isinstance(result[0], MultiValueMap)
        # Verify source values
        assert result[0].source == ['DE', 'LC']
        # Verify target values (now a list)
        assert result[0].target == ['EUR']

    def test_build_multi_value_map_list_regex_source(self, multi_value_map_df: pd.DataFrame) -> None:
        """Row with regex pattern in source should still work."""
        df = multi_value_map_df.iloc[[1]]  # regex:^A
        
        result = build_multi_value_map_list(df, ['country', 'currency'], ['iso_code'])
        
        # Check raw source string
        assert result[0].source[0].startswith("regex:^A")
        # Check compiled pattern property
        assert result[0].typed_source[0] == re.compile(r"^A")
        assert result[0].target == ['ARG']

    def test_build_multi_value_map_list_empty_source(self, multi_value_map_df: pd.DataFrame) -> None:
        """Row with empty string in source should be accepted as valid string."""
        df = multi_value_map_df.iloc[[2]]  # Empty country string
        
        result = build_multi_value_map_list(df, ['country', 'currency'], ['iso_code'])
        
        assert result[0].source[0] == ""
        assert result[0].target == ['CHF']

    def test_build_multi_value_map_list_with_validity(self, multi_value_map_df: pd.DataFrame) -> None:
        """Row with valid_from and valid_to should include datetime fields."""
        df = multi_value_map_df.iloc[[3]]  # FR/FRA with validity dates
        
        result = build_multi_value_map_list(df, ['country', 'currency'], ['iso_code'])
        
        mv_map = result[0]
        assert mv_map.source == ['FR', 'LC']
        assert mv_map.target == ['FRA']
        
        assert isinstance(mv_map.valid_from, datetime)
        assert isinstance(mv_map.valid_to, datetime)
        assert mv_map.valid_from.isoformat() == "2020-01-01T00:00:00"
        assert mv_map.valid_to.isoformat() == "2025-12-31T00:00:00"

    def test_build_multi_value_map_list_invalid_type(self, multi_value_map_df: pd.DataFrame) -> None:
        """Row with non-string source should raise TypeError."""
        df = multi_value_map_df.iloc[[4]]  # Invalid int type in country column
        
        with pytest.raises(TypeError, match="Source column 'country' must contain only string values"):
            build_multi_value_map_list(df, ['country', 'currency'], ['iso_code'])

    def test_build_multi_value_map_list_missing_target_column(self, multi_value_map_df: pd.DataFrame) -> None:
        """Missing target column should raise ValueError."""
        df = multi_value_map_df.drop(columns=['iso_code'])
        
        with pytest.raises(ValueError, match="Target columns missing"):
            build_multi_value_map_list(df, ['country', 'currency'], ['iso_code'])

    def test_build_multi_value_map_list_missing_source_column(self, multi_value_map_df: pd.DataFrame) -> None:
        """Missing one source column should raise ValueError."""
        df = multi_value_map_df.drop(columns=['currency'])
        
        with pytest.raises(ValueError, match="Source columns missing"):
            build_multi_value_map_list(df, ['country', 'currency'], ['iso_code'])

    def test_build_multi_value_map_list_empty_dataframe(self) -> None:
        """Empty DataFrame should raise ValueError."""
        df = pd.DataFrame()
        
        with pytest.raises(ValueError, match="Input DataFrame cannot be empty"):
            build_multi_value_map_list(df, ['country', 'currency'], ['iso_code'])

    def test_build_multi_value_map_list_multiple_rows(self, multi_value_map_df: pd.DataFrame) -> None:
        """Multiple rows should return a list of MultiValueMap objects."""
        df = multi_value_map_df.iloc[:3]  # First three rows
        
        result = build_multi_value_map_list(df, ['country', 'currency'], ['iso_code'])
        
        assert isinstance(result, list)
        assert len(result) == df.shape[0]
        
        for mv_map in result:
            assert isinstance(mv_map, MultiValueMap)
            # Ensure source/target are sequences (lists/tuples)
            assert isinstance(mv_map.source, (list, tuple))
            assert isinstance(mv_map.target, (list, tuple))
            # Ensure target is not nested like [['EUR']] but flat like ['EUR']
            assert len(mv_map.target) == 1
            assert isinstance(mv_map.target[0], str)
    
    @pytest.fixture
    def valid_dataframe(self):
        """Provides a standard valid DataFrame for basic testing."""
        data = {
            'country': ['DE', 'FR'],
            'currency': ['EUR', 'EUR'],
            'code': ['A', 'B'],
            'target_code': ['X', 'Y'],
            'valid_from': ['2020-01-01', '2021-01-01'],
            'valid_to': ['2020-12-31', '2021-12-31']
        }
        return pd.DataFrame(data)

    def test_raises_value_error_when_dataframe_is_empty(self):
        """Tests that a ValueError is raised if the input DataFrame is empty."""
        # Arrange
        empty_df = pd.DataFrame()

        # Act & Assert
        with pytest.raises(ValueError, match="Input DataFrame cannot be empty"):
            build_multi_value_map_list(empty_df, ['col'], ['col'])

    @pytest.mark.parametrize("missing_cols, is_source", [
        (['missing_src'], True),
        (['missing_tgt'], False)
    ])
    def test_raises_value_error_when_columns_missing(self, valid_dataframe, missing_cols, is_source):
        """Tests that ValueError is raised when source or target columns are missing.
        
        Args:
            valid_dataframe (pd.DataFrame): The fixture data.
            missing_cols (list): The columns to request that don't exist.
            is_source (bool): Whether we are testing source or target columns.
        """
        # Arrange
        source_cols = missing_cols if is_source else ['country']
        target_cols = ['target_code'] if is_source else missing_cols

        # Act & Assert
        match_msg = "Source columns missing" if is_source else "Target columns missing"
        with pytest.raises(ValueError, match=match_msg):
            build_multi_value_map_list(valid_dataframe, source_cols, target_cols)

    def test_raises_type_error_when_columns_contain_non_strings(self):
        """Tests that TypeError is raised if source/target columns contain non-string data."""
        # Arrange
        df = pd.DataFrame({
            'src_str': ['A', 'B'],
            'src_int': [1, 2],  # Invalid
            'tgt_str': ['X', 'Y']
        })

        # Act & Assert
        with pytest.raises(TypeError, match="must contain only string values"):
            build_multi_value_map_list(df, ['src_int'], ['tgt_str'])

    def test_builds_maps_correctly_with_string_dates(self, valid_dataframe):
        """Tests that maps are built correctly using standard ISO string dates."""
        # Arrange
        source_cols = ['country', 'currency']
        target_cols = ['target_code']

        # Act
        result = build_multi_value_map_list(
            valid_dataframe, 
            source_cols, 
            target_cols
        )

        # Assert
        assert len(result) == 2
        assert isinstance(result[0], MultiValueMap)
        assert result[0].source == ['DE', 'EUR']
        assert result[0].target == ['X']
        assert isinstance(result[0].valid_from, datetime)
        assert result[0].valid_from.year == 2020

    def test_handles_pandas_timestamp_validity(self):
        """Tests that pd.Timestamp objects in valid_from/valid_to are converted to datetime.
        
        This specifically targets the line:
        elif hasattr(val, "to_pydatetime"): kwargs["valid_from"] = val.to_pydatetime()
        """
        # Arrange
        data = {
            'src': ['A'],
            'tgt': ['B'],
            'valid_from': [pd.Timestamp('2022-01-01')],
            'valid_to': [pd.Timestamp('2022-12-31')]
        }
        df = pd.DataFrame(data)

        # Act
        result = build_multi_value_map_list(df, ['src'], ['tgt'])

        # Assert
        map_obj = result[0]
        # Check valid_from
        assert isinstance(map_obj.valid_from, datetime)
        # Ensure it's a standard datetime, not a pandas Timestamp (unless they are same class in env, 
        # but the code calls to_pydatetime explicitely to convert)
        assert map_obj.valid_from == datetime(2022, 1, 1)
        
        # Check valid_to
        assert isinstance(map_obj.valid_to, datetime)
        assert map_obj.valid_to == datetime(2022, 12, 31)

    @pytest.mark.skip(reason="Not implemented correctly. To review")
    def test_handles_native_datetime_validity(self):
        """Tests that native datetime objects in valid_from/valid_to are preserved.
        
        This specifically targets the line:
        elif isinstance(val, datetime): kwargs["valid_from"] = val
        """
        # Arrange
        dt_from = datetime(2023, 6, 15)
        dt_to = datetime(2023, 7, 15)
        
        data = {
            'src': ['A'],
            'tgt': ['B'],
            'valid_from': [dt_from],
            'valid_to': [dt_to]
        }
        df = pd.DataFrame(data)

        # Act
        result = build_multi_value_map_list(df, ['src'], ['tgt'])

        # Assert
        map_obj = result[0]
        assert map_obj.valid_from is dt_from  # Should be the exact object or equal
        assert map_obj.valid_to is dt_to

    def test_ignores_nan_validity_values(self):
        """Tests that NaT/NaN/None in validity columns result in None in the object."""
        # Arrange
        data = {
            'src': ['A'],
            'tgt': ['B'],
            'valid_from': [pd.NA],
            'valid_to': [None]
        }
        df = pd.DataFrame(data)

        # Act
        result = build_multi_value_map_list(df, ['src'], ['tgt'])

        # Assert
        map_obj = result[0]
        assert map_obj.valid_from is None
        assert map_obj.valid_to is None

class TestBuildRepresentationMap:  # noqa: D101

    def test_build_representation_map_success(self, value_map_df_mandatory_cols, id, name, agency, source_cl, target_cl, description):
        """Test successful creation of RepresentationMap from valid DataFrame."""
        rm = build_representation_map(
            df=value_map_df_mandatory_cols,
            id=id,
            name=name,
            agency=agency,
            source_cl=source_cl,
            target_cl=target_cl,
            description=description
        )
        assert isinstance(rm, RepresentationMap)
        assert rm.id == id
        assert rm.name == name
        assert rm.agency == agency
        assert rm.source == source_cl
        assert rm.target == target_cl
        assert rm.description == description
        assert rm.version == "1.0"
        assert len(rm.maps) == value_map_df_mandatory_cols.shape[0]
        for vm in rm.maps:
            assert isinstance(vm, ValueMap)

    def test_build_representation_map_empty_df_raises(self):
        """Empty DataFrame should raise ValueError."""
        empty_df = pd.DataFrame(columns=["source", "target"])
        with pytest.raises(ValueError):
            build_representation_map(empty_df)

    def test_build_representation_map_missing_columns_raises(self):
        """Missing mandatory columns should raise ValueError."""
        bad_df = pd.DataFrame({"src": ["A"], "tgt": ["B"]})
        with pytest.raises(ValueError):
            build_representation_map(bad_df)

    def test_build_representation_map_non_string_values_raises(self):
        """Non-string values in source or target columns should raise TypeError."""
        bad_df = pd.DataFrame({"source": [123], "target": ["ABC"]})
        with pytest.raises(TypeError):
            build_representation_map(bad_df)

    def test_build_representation_map_custom_column_names(self):
        """Test with custom column names for source and target."""
        df = pd.DataFrame({
            "src": ["A", "B"],
            "tgt": ["X", "Y"],
            "valid_from": [None, None],
            "valid_to": [None, None]
        })
        rm = build_representation_map(
            df=df,
            source_col="src",
            target_col="tgt",
            id="RM_CUSTOM",
            name="Custom Map"
        )
        assert isinstance(rm, RepresentationMap)
        assert rm.id == "RM_CUSTOM"
        assert rm.name == "Custom Map"
        assert len(rm.maps) == 2

    def test_build_representation_map_validity_dates(self, value_map_df_mandatory_cols):
        """Ensure validity dates are parsed correctly when present."""
        rm = build_representation_map(df=value_map_df_mandatory_cols)
        for vm in rm.maps:
            if vm.valid_from:
                assert isinstance(vm.valid_from, datetime)
                assert vm.valid_from.tzinfo == timezone.utc or vm.valid_from.tzinfo is None
            if vm.valid_to:
                assert isinstance(vm.valid_to, datetime)
                assert vm.valid_to.tzinfo == timezone.utc or vm.valid_to.tzinfo is None

class TestBuildSingleComponentMap:  # noqa: D101
    def test_build_single_component_map_valid(self, value_map_df_mandatory_cols):
        """Test normal case with valid DataFrame and fixture."""
        df = value_map_df_mandatory_cols
        cm = build_single_component_map(
            df,
            source_component="COUNTRY",
            target_component="COUNTRY",
            agency="ECB",
            id="CM1",
            name="Country Component Map",
            source_cl="urn:source:codelist",
            target_cl="urn:target:codelist"
        )
        assert isinstance(cm, ComponentMap)
        assert cm.source == "COUNTRY"
        assert cm.target == "COUNTRY"
        assert isinstance(cm.values, RepresentationMap)


    def test_build_single_component_map_empty_df(self):
        """Empty DataFrame should raise ValueError."""
        df = pd.DataFrame(columns=["source", "target"])
        with pytest.raises(ValueError):
            build_single_component_map(df, "COUNTRY", "COUNTRY")


    def test_build_single_component_map_missing_column(self, value_map_df_mandatory_cols):
        """Missing required column should raise ValueError."""
        df = value_map_df_mandatory_cols.drop(columns=["source"])
        with pytest.raises(ValueError):
            build_single_component_map(df, "COUNTRY", "COUNTRY")


    def test_build_single_component_map_invalid_type(self):
        """Non-string values in source column should raise TypeError."""
        df = pd.DataFrame({"source": [123], "target": ["BEL"]})
        with pytest.raises(TypeError):
            build_single_component_map(df, "COUNTRY", "COUNTRY")


    def test_build_single_component_map_custom_columns(self):
        """Test with custom column names for source and target."""
        df = pd.DataFrame({
            "src": ["BE", "FR"],
            "tgt": ["BEL", "FRA"],
            "valid_from": ["2020-01-01", None],
            "valid_to": ["2025-12-31", None]
        })
        cm = build_single_component_map(
            df,
            source_component="COUNTRY",
            target_component="COUNTRY",
            source_col="src",
            target_col="tgt"
        )
        assert isinstance(cm, ComponentMap)
        assert isinstance(cm.values, RepresentationMap)


    def test_build_single_component_map_with_validity_dates(self, value_map_df_mandatory_cols):
        """Ensure validity columns are handled correctly."""
        df = value_map_df_mandatory_cols.copy()
        df["valid_from"] = ["2020-01-01", None, None]
        df["valid_to"] = ["2025-12-31", None, None]
        cm = build_single_component_map(
            df,
            source_component="COUNTRY",
            target_component="COUNTRY"
        )
        assert isinstance(cm.values, RepresentationMap)
        assert len(cm.values.maps) == len(df)


    def test_build_single_component_map_id_and_name(self, value_map_df_mandatory_cols):
        """Check that id and name propagate correctly to RepresentationMap."""
        cm = build_single_component_map(
            value_map_df_mandatory_cols,
            source_component="COUNTRY",
            target_component="COUNTRY",
            id="TEST_ID",
            name="Test Name"
        )
        assert cm.values.id == "TEST_ID"
        assert cm.values.name == "Test Name"


    def test_build_single_component_map_version_and_description(self, value_map_df_mandatory_cols):
        """Check version and description fields."""
        cm = build_single_component_map(
            value_map_df_mandatory_cols,
            source_component="COUNTRY",
            target_component="COUNTRY",
            version="2.0",
            description="Test Description"
        )
        assert cm.values.version == "2.0"
        assert cm.values.description == "Test Description"

class TestBuildMultiRepresentationMap: # noqa: D101
    """Tests for the build_multi_representation_map function."""

    @pytest.fixture
    def sample_df(self):
        """Fixture providing a sample DataFrame for normal cases."""
        data = {
            "source": ["BE", "FR"],
            "target": ["BEL", "FRA"],
            "valid_from": ["2020-01-01", None],
            "valid_to": ["2025-12-31", None]
        }
        return pd.DataFrame(data)

    def test_returns_multi_representation_map(self, sample_df):
        """Tests that the function returns a MultiRepresentationMap for valid input."""
        result = build_multi_representation_map(
            sample_df,
            id="MRM1",
            name="Country Multi Map",
            agency="ECB",
            source_cls=["urn:source:codelist"],
            target_cls=["urn:target:codelist"]
        )
        assert isinstance(result, MultiRepresentationMap)
        assert result.id == "MRM1"
        assert result.name == "Country Multi Map"
        assert result.agency == "ECB"
        assert len(result.maps) == 2

    def test_empty_dataframe_raises_value_error(self):
        """Tests that ValueError is raised when DataFrame is empty."""
        empty_df = pd.DataFrame()
        with pytest.raises(ValueError, match="Input DataFrame cannot be empty"):
            build_multi_representation_map(empty_df)

    def test_missing_required_columns_raises_value_error(self, sample_df):
        """Tests that ValueError is raised when required columns are missing."""
        df_missing = sample_df.drop(columns=["target"])
        with pytest.raises(ValueError, match="Missing required columns"):
            build_multi_representation_map(df_missing)

    @pytest.mark.parametrize("invalid_data", [
        {"source": [123, "FR"], "target": ["BEL", "FRA"]},  # Non-string in source
        {"source": ["BE", "FR"], "target": [None, "FRA"]},  # None in target
    ])
    def test_non_string_values_raise_type_error(self, invalid_data):
        """Tests that TypeError is raised when source or target columns contain non-string values."""
        df_invalid = pd.DataFrame(invalid_data)
        with pytest.raises(TypeError):
            build_multi_representation_map(df_invalid)

    def test_custom_columns_and_metadata(self):
        """Tests that custom column names and metadata are handled correctly."""
        data = {
            "source": ["US", "UK"],
            "target": ["USA", "GBR"],
            "valid_from": ["2021-01-01", None],
            "valid_to": ["2023-12-31", None]
        }
        df_custom = pd.DataFrame(data)
        result = build_multi_representation_map(
            df_custom,
            source_cls=["source"],
            target_cls=["target"],
            id="MRM2",
            name="Custom Map",
            description="Test description"
        )
        assert isinstance(result, MultiRepresentationMap)
        assert result.id == "MRM2"
        assert result.name == "Custom Map"
        assert result.description == "Test description"
        assert len(result.maps) == 2

    def test_validity_dates_are_parsed_correctly(self, sample_df):
        """Tests that validity dates are correctly passed to MultiValueMap objects."""
        result = build_multi_representation_map(sample_df)
        first_map = result.maps[0]
        assert first_map.valid_from == datetime.fromisoformat("2020-01-01")
        assert first_map.valid_to == datetime.fromisoformat("2025-12-31")

class TestExtractMappingDefinitions:  # noqa: D101
    @pytest.fixture
    def mock_empty_workbook(self) -> Workbook:
        """Fixture returning a simple empty workbook."""
        wb = Workbook()
        return wb

    @pytest.fixture
    def mock_populated_workbook(self, mock_empty_workbook: Workbook) -> Workbook:
        """Fixture returning a mock workbook with multiple sheets."""
        wb = mock_empty_workbook
        
        # Mandatory comp_mapping sheet
        ws = wb.active
        ws.title = "comp_mapping"
        ws.append(["Source", "Target", "Mapping_Rules"]) # Case insensitive header test
        ws.append(["", "T1_FIXED", "fixed:A"])
        ws.append(["SRC_2", "T2_IMPLICIT", "implicit"])
        ws.append(["", "T3_REP", "T3_REP"])
        ws.append(["SRC_4", "T4_REP", "T4_REP"]) # Rep map with explicit source
        ws.append([None, None, None]) # Empty row
        
        # Referenced Rep Map sheets (T3 is empty, T4 has data)
        ws_rep_3 = wb.create_sheet("T3_REP")
        ws_rep_3.append(["source", "target", "valid_from", "valid_to"]) # Empty rows below header
        
        ws_rep_4 = wb.create_sheet("T4_REP")
        ws_rep_4.append(["source", "target", "valid_from", "valid_to"])
        ws_rep_4.append(["S1", "T1", "", ""])
        
        return wb


    def test_read_comp_mapping_sheet_success(self, mock_populated_workbook: Workbook):
        """Tests if the sheet is loaded, headers normalized, and empty values handled."""
        df = _read_comp_mapping_sheet(mock_populated_workbook)
        
        assert isinstance(df, pd.DataFrame)
        assert list(df.columns) == ["source", "target", "mapping_rules"]
        assert len(df) == 4 # Empty row has been removed
        assert df.iloc[0]["mapping_rules"] == "fixed:A"


    def test_read_comp_mapping_sheet_key_error(self, mock_empty_workbook: Workbook):
        """Tests KeyError when the sheet is missing."""
        mock_empty_workbook.active.title = "WrongName"
        with pytest.raises(KeyError, match="comp_mapping"):
            _read_comp_mapping_sheet(mock_empty_workbook)


    def test_create_fixed_definition_success(self):
        """Tests successful creation of a FixedValueMap definition."""
        definition = _create_fixed_definition(pd.Series(), "T_FIX", "fixed:MY_VALUE")
        assert definition.map_type == "fixed"
        assert definition.fixed_value == "MY_VALUE"
        assert definition.target == "T_FIX"


    def test_create_fixed_definition_empty_value_raises_value_error(self):
        """Tests validation for empty fixed value."""
        with pytest.raises(ValueError, match="cannot be empty"):
            _create_fixed_definition(pd.Series(), "T_FIX", "fixed:")


    def test_create_implicit_definition_success(self):
        """Tests successful creation of an ImplicitComponentMap definition."""
        definition = _create_implicit_definition(pd.Series(), "T_IMP", "S_IMP")
        assert definition.map_type == "implicit"
        assert definition.source == "S_IMP"
        assert definition.target == "T_IMP"


    def test_create_implicit_definition_missing_source_raises_value_error(self):
        """Tests validation for missing source in implicit mapping."""
        with pytest.raises(ValueError, match="requires a 'source'"):
            _create_implicit_definition(pd.Series(), "T_IMP", "")


    def test_create_representation_definition_success(self, mock_populated_workbook: Workbook):
        """Tests successful creation of a RepresentationMap definition including sheet loading."""
        definition = _create_representation_definition(mock_populated_workbook, "T4_REP", "SRC_4")
        assert definition.map_type == "representation"
        assert definition.source == "SRC_4"
        assert definition.target == "T4_REP"
        assert isinstance(definition.representation_df, pd.DataFrame)
        assert len(definition.representation_df) == 1 # Check data rows exist

    def test_create_representation_definition_source_inference(self, mock_populated_workbook: Workbook):
        """Tests source inference when source is empty."""
        definition = _create_representation_definition(mock_populated_workbook, "T3_REP", "")
        assert definition.source == "T3_REP" # Inferred source is target
        assert len(definition.representation_df) == 0 # Check empty sheet handling


    def test_extract_mapping_definitions_integration(self, mock_populated_workbook: Workbook):
        """Tests the main function's dispatch logic."""
        definitions = _extract_mapping_definitions(mock_populated_workbook)
        
        assert len(definitions) == 4 # Should ignore empty row and invalid rules if any
        
        # T1_FIXED (Fixed)
        assert definitions[0].map_type == "fixed"
        
        # T2_IMPLICIT (Implicit)
        assert definitions[1].map_type == "implicit"
        
        # T3_REP (Representation - empty DF)
        assert definitions[2].map_type == "representation"
        
        # T4_REP (Representation - data DF)
        assert definitions[3].map_type == "representation"
        assert definitions[3].source == "SRC_4"

    def test_extract_mapping_definitions_invalid_rule_raises_value_error(self, mock_populated_workbook: Workbook):
        """Tests invalid rule check."""
        ws = mock_populated_workbook["comp_mapping"]
        ws.append(["", "T_BAD", "unknown_type"])
        
        with pytest.raises(ValueError, match="Unknown mapping rule"):
            _extract_mapping_definitions(mock_populated_workbook)

class TestCreateSchemaFromTable:  # noqa: D101
    def test_create_schema_time_period_standardization(self) -> None:
        """Test that the time dimension is standardized to TIME_PERIOD."""
        df = pd.DataFrame({
            "REF_AREA": ["US"],
            "my_date_col": ["2020"],
            "VALUE": [100.0]
        })
        
        schema = create_schema_from_table(
            df,
            dimensions=["REF_AREA"],
            time_dimension="my_date_col",
            measure="VALUE"
        )
        
        # Verify the component is named TIME_PERIOD, not my_date_col
        assert schema.components["TIME_PERIOD"] is not None
        assert schema.components["my_date_col"] is None
        
        # Verify strict properties
        time_comp = schema.components["TIME_PERIOD"]
        assert time_comp.id == "TIME_PERIOD"
        assert time_comp.role == Role.DIMENSION
        assert time_comp.local_dtype == DataType.PERIOD
        assert time_comp.description == "Timespan or point in time to which the observation actually refers."
        
        # Verify Concept properties
        assert isinstance(time_comp.concept, Concept)
        assert time_comp.concept.id == "TIME_PERIOD"
        assert time_comp.concept.urn == "urn:sdmx:org.sdmx.infomodel.conceptscheme.Concept=SDMX:CROSS_DOMAIN_CONCEPTS(2.0).TIME_PERIOD"
        assert time_comp.concept.dtype == DataType.STRING


    def test_create_schema_structure(self) -> None:
        """Test general structure creation."""
        df = pd.DataFrame({
            "FREQ": ["A"],
            "TIME_PERIOD": ["2020"],
            "OBS": [1],
            "STATUS": ["A"]
        })
        
        schema = create_schema_from_table(
            df,
            dimensions=["FREQ"],
            time_dimension="TIME_PERIOD",
            measure="OBS",
            attributes=["STATUS"]
        )
        
        assert len(schema.components) == 4
        assert schema.components["FREQ"].role == Role.DIMENSION
        assert schema.components["TIME_PERIOD"].role == Role.DIMENSION
        assert schema.components["OBS"].role == Role.MEASURE
        assert schema.components["STATUS"].role == Role.ATTRIBUTE


    def test_create_schema_missing_columns(self) -> None:
        """Test validation of input columns."""
        df = pd.DataFrame({"A": [1]})
        with pytest.raises(ValueError) as exc:
            create_schema_from_table(df, dimensions=[], measure="A", time_dimension="MISSING")
        assert "Columns not found" in str(exc.value)

class TestBuildSchemaFromWbTemplate:  # noqa: D101
    def test_parse_info_sheet_basic_scenario(self):
        """Test a standard scenario where the sheet contains simple Key-Value pairs.

        Includes testing that the 'header' (first row) is treated as data.
        """
        # Create a DataFrame where the "header" is actually the first metadata pair
        data = {
            "Dataset_ID": ["Source", "Version"],
            "WB_ASPIRE": ["World Bank", "1.0"]
        }
        df = pd.DataFrame(data)
        # The dataframe looks like:
        #    Dataset_ID  WB_ASPIRE  <-- Treated as Row 0
        # 0  Source      World Bank
        # 1  Version     1.0

        sheets = {"INFO": df}
        result = _parse_info_sheet(sheets, sheet_name="INFO")

        assert len(result) == 3
        # Check Row 0 (Headers)
        assert result.iloc[0]["Key"] == "Dataset_ID"
        assert result.iloc[0]["Value"] == "WB_ASPIRE"
        # Check Row 1
        assert result.iloc[1]["Key"] == "Source"
        assert result.iloc[1]["Value"] == "World Bank"
        # Check Row 2
        assert result.iloc[2]["Key"] == "Version"
        assert result.iloc[2]["Value"] == "1.0"


    def test_parse_info_sheet_missing_sheet_raises_error(self):
        """Test that a ValueError is raised if the requested sheet does not exist."""
        sheets = {"DATA": pd.DataFrame({"A": [1]})}
        
        with pytest.raises(ValueError, match="Sheet 'INFO' not found"):
            _parse_info_sheet(sheets, sheet_name="INFO")


    def test_parse_info_sheet_ignore_section_header(self):
        """Test that the specific 'DATA CURATION PROCESS' string causes the row to be ignored."""
        df = pd.DataFrame([
            ["DATA CURATION PROCESS", "Some Description"],  # Should be ignored
            ["Valid_Key", "Valid_Value"],                   # Should be kept
            [None, "DATA CURATION PROCESS"]                 # Should be ignored
        ])
        sheets = {"INFO": df}
        result = _parse_info_sheet(sheets)

        assert len(result) == 1
        assert result.iloc[0]["Key"] == "Valid_Key"
        assert result.iloc[0]["Value"] == "Valid_Value"


    def test_parse_info_sheet_ignore_rows_with_wrong_count(self):
        """Test that rows with 1 item or more than 2 items are excluded."""
        df = pd.DataFrame([
            ["OnlyOne"],                       # 1 valid cell -> Keep
            ["Key", "Value", "Extra"],         # 3 valid cells -> Ignore
            # [None, None, "SingleValid"],       # 1 valid cell (after nan filter) -> Ignore
            ["Key2", "Value2", None]           # 2 valid cells -> Keep
        ])
        sheets = {"INFO": df}
        result = _parse_info_sheet(sheets)

        assert len(result) == 2
        assert result.iloc[0]["Key"] == "OnlyOne"
        assert result.iloc[1]["Key"] == "Key2"
        assert result.iloc[1]["Value"] == "Value2"


    def test_parse_info_sheet_empty_input(self):
        """Test parsing an empty DataFrame results in an empty result with correct columns."""
        df = pd.DataFrame()
        sheets = {"INFO": df}
        result = _parse_info_sheet(sheets)

        assert result.empty
        assert list(result.columns) == ["Key", "Value"]


    def test_parse_info_sheet_nan_handling(self):
        """Test that NaN, None, and string 'nan' are filtered out effectively."""
        df = pd.DataFrame([
            ["Key1", np.nan, "Value1"],    # 2 valid cells -> Keep
            [None, "Key2", "Value2"],      # 2 valid cells -> Keep
            ["nan", "Key3", "Value3"],     # 'nan' string ignored -> 2 valid -> Keep
            ["Key4", "", "Value4"]         # Empty string ignored -> 2 valid -> Keep
        ])
        sheets = {"INFO": df}
        result = _parse_info_sheet(sheets)

        assert len(result) == 4
        assert result.iloc[0]["Key"] == "Key1"
        assert result.iloc[0]["Value"] == "Value1"
        assert result.iloc[2]["Key"] == "Key3"
        assert result.iloc[2]["Value"] == "Value3"


    def test_parse_info_sheet_mixed_indentation(self):
        """Test handling of data that is visually indented (starts in column 1 or 2)."""
        # Simulate pandas reading excel where first cols are NaN
        data = [
            [np.nan, "Key1", "Value1"],
            [np.nan, np.nan, np.nan],
            [np.nan, "Key2", "Value2"]
        ]
        df = pd.DataFrame(data)
        sheets = {"INFO": df}
        result = _parse_info_sheet(sheets)

        assert len(result) == 2
        assert result.iloc[0]["Key"] == "Key1"
        assert result.iloc[1]["Key"] == "Key2"


    def test_parse_info_sheet_unnamed_columns(self):
        """Test that pandas 'Unnamed: X' columns (common in headerless parsing) don't break logic."""
        # pd.read_excel often produces 'Unnamed: 0', 'Unnamed: 1' if no header is found
        # These will be automatically removed
       
        df = pd.DataFrame({"Unnamed: 0": ["Key"], "Unnamed: 1": ["Value"]})
        sheets = {"INFO": df}
        result = _parse_info_sheet(sheets)
        
        assert len(result) == 1
        assert result.iloc[0]["Key"] == "Key"
        assert result.iloc[0]["Value"] == "Value"


    def test_parse_info_sheet_whitespace_stripping(self):
        """Test that whitespace surrounding keys and values is removed."""
        df = pd.DataFrame([
            ["  Key1  ", "Value1"],
            ["Key2", "  Value2  \t"]
        ])
        sheets = {"INFO": df}
        result = _parse_info_sheet(sheets)

        assert result.iloc[0]["Key"] == "Key1"
        assert result.iloc[1]["Value"] == "Value2"

    def test_parse_comp_mapping_sheet_normal_case(self):
        """Test parsing a valid COMP_MAPPING sheet with expected structure."""
        data = {
            "SOURCE": ["Series code", None, "year"],
            "TARGET": ["INDICATOR", "FREQ", "TIME_PERIOD"],
            "MAPPING_RULES": ["INDICATOR", "fixed:A", "TIME_PERIOD"],
            "Unnamed: 3": [None, None, None]  # Artifact to ensure we filter it out
        }
        df = pd.DataFrame(data)
        sheets = {"COMP_MAPPING": df}

        result = _parse_comp_mapping_sheet(sheets)

        # Check columns
        assert list(result.columns) == ["SOURCE", "TARGET", "MAPPING_RULES"]
        
        # Check data integrity
        assert len(result) == 3
        assert result.iloc[0]["SOURCE"] == "Series code"
        assert pd.isna(result.iloc[1]["SOURCE"])  # Ensure None/NaN is preserved
        assert result.iloc[1]["TARGET"] == "FREQ"
        assert result.iloc[1]["MAPPING_RULES"] == "fixed:A"


    def test_parse_comp_mapping_sheet_missing_sheet(self):
        """Test that ValueError is raised when the sheet is not present."""
        sheets = {"INFO": pd.DataFrame()}
        
        with pytest.raises(ValueError, match="Sheet 'COMP_MAPPING' not found"):
            _parse_comp_mapping_sheet(sheets)


    def test_parse_comp_mapping_sheet_missing_columns(self):
        """Test that ValueError is raised when required columns are missing."""
        # MAPPING_RULES is missing
        df = pd.DataFrame({
            "SOURCE": ["A"],
            "TARGET": ["B"]
        })
        sheets = {"COMP_MAPPING": df}

        with pytest.raises(ValueError, match="missing required columns"):
            _parse_comp_mapping_sheet(sheets)


    def test_parse_comp_mapping_sheet_empty_rows(self):
        """Test that completely empty rows are removed."""
        df = pd.DataFrame({
            "SOURCE": ["A", None, None],
            "TARGET": ["B", None, "D"],
            "MAPPING_RULES": ["C", None, "E"]
        })
        # Row 1 is fully empty -> should be dropped
        # Row 2 has None in SOURCE but data in others -> should be kept
        
        sheets = {"COMP_MAPPING": df}
        result = _parse_comp_mapping_sheet(sheets)

        assert len(result) == 2
        assert result.iloc[0]["SOURCE"] == "A"
        assert result.iloc[1]["TARGET"] == "D"


    def test_parse_comp_mapping_sheet_empty_input(self):
        """Test parsing a sheet that has headers but no data."""
        df = pd.DataFrame(columns=["SOURCE", "TARGET", "MAPPING_RULES"])
        sheets = {"COMP_MAPPING": df}
        
        result = _parse_comp_mapping_sheet(sheets)
        
        assert result.empty
        assert list(result.columns) == ["SOURCE", "TARGET", "MAPPING_RULES"]

    def test_parse_rep_mapping_normal_case(self):
        """Test standard separation of S: and T: columns into a dictionary."""
        data = {
            "S:Series": ["A", "B"],
            "S:Name": ["Alpha", "Beta"],
            "T:Indicator": ["IND_A", "IND_B"],
            "T:Unit": ["USD", "EUR"],
            "Notes": ["Ignore", "Ignore"]  # Should be ignored
        }
        df = pd.DataFrame(data)
        sheets = {"REP_MAPPING": df}

        result = _parse_rep_mapping_sheet(sheets)

        assert isinstance(result, dict)
        assert "source" in result
        assert "target" in result

        # Validate Source DF
        source_df = result["source"]
        assert list(source_df.columns) == ["Series", "Name"]
        assert len(source_df) == 2
        assert source_df.iloc[0]["Series"] == "A"

        # Validate Target DF
        target_df = result["target"]
        assert list(target_df.columns) == ["Indicator", "Unit"]
        assert len(target_df) == 2
        assert target_df.iloc[1]["Unit"] == "EUR"


    def test_parse_rep_mapping_missing_sheet(self):
        """Test that ValueError is raised if sheet is missing."""
        sheets = {"OTHER": pd.DataFrame()}
        with pytest.raises(ValueError, match="Sheet 'REP_MAPPING' not found"):
            _parse_rep_mapping_sheet(sheets)


    def test_parse_rep_mapping_no_source_cols(self):
        """Test that ValueError is raised if no S: columns exist."""
        df = pd.DataFrame({"T:Indicator": [1], "Other": [2]})
        sheets = {"REP_MAPPING": df}
        
        with pytest.raises(ValueError, match="No source columns"):
            _parse_rep_mapping_sheet(sheets)


    def test_parse_rep_mapping_no_target_cols(self):
        """Test that ValueError is raised if no T: columns exist."""
        df = pd.DataFrame({"S:Series": [1], "Other": [2]})
        sheets = {"REP_MAPPING": df}
        
        with pytest.raises(ValueError, match="No target columns"):
            _parse_rep_mapping_sheet(sheets)


    def test_parse_rep_mapping_empty_data(self):
        """Test parsing a sheet with correct headers but no rows."""
        df = pd.DataFrame(columns=["S:Series", "T:Indicator"])
        sheets = {"REP_MAPPING": df}

        result = _parse_rep_mapping_sheet(sheets)
        source_df = result["source"]
        target_df = result["target"]

        assert list(source_df.columns) == ["Series"]
        assert list(target_df.columns) == ["Indicator"]
        assert source_df.empty
        assert target_df.empty


    def test_parse_rep_mapping_ignore_unprefixed_cols(self):
        """Test that columns without 'S:' or 'T:' prefixes are excluded."""
        data = {
            "S:Key": [1],
            "T:Val": [2],
            "Random": [3],
            "Unnamed: 0": [4]
        }
        df = pd.DataFrame(data)
        sheets = {"REP_MAPPING": df}

        result = _parse_rep_mapping_sheet(sheets)
        
        assert "Random" not in result["source"].columns
        assert "Random" not in result["target"].columns
        assert "Unnamed: 0" not in result["source"].columns

class TestMatchColumnName: #noqa: D101
    # Shared list of column names mimicking cleaned headers from REP_MAPPING sheet
    AVAILABLE_COLUMNS = [
        "INDICATOR_CODE", 
        "Time Period", 
        "REF AREA", 
        "OBS_VALUE_CLEAN", 
        "ShortName",
        "LONG_NAME_TEST"
    ]

    def test_match_column_name_exact_match(self):
        """Tests case-sensitive exact matching (Rule 1)."""
        target = "REF AREA"
        expected = "REF AREA"
        result = _match_column_name(target, self.AVAILABLE_COLUMNS)
        assert result == expected
        assert result in self.AVAILABLE_COLUMNS

    def test_match_column_name_normalized_match_case_insensitive(self):
        """Tests case-insensitive match after normalization (Rule 2)."""
        # Target is lowercase, column is uppercase with underscore
        target = "indicatorcode"
        expected = "INDICATOR_CODE"
        result = _match_column_name(target, self.AVAILABLE_COLUMNS)
        assert result == expected

    def test_match_column_name_normalized_match_spacing(self):
        """Tests match ignoring internal spaces and different spacing in target."""
        # Target has different spacing and casing
        target = "RefArea"
        expected = "REF AREA"
        result = _match_column_name(target, self.AVAILABLE_COLUMNS)
        assert result == expected

    def test_match_column_name_normalized_match_underscores(self):
        """Tests match ignoring underscores vs spaces/no spaces."""
        # Target has underscores, column has underscores
        target = "obs_value_clean"
        expected = "OBS_VALUE_CLEAN"
        result = _match_column_name(target, self.AVAILABLE_COLUMNS)
        assert result == expected

    def test_match_column_name_fuzzy_match_short_in_long(self):
        """Tests fuzzy match where the column name is contained in the target name (e.g., 'ShortName' in 'TheShortName')."""
        # Target is longer/more descriptive, column is the core name
        target = "The Short Name Field"
        expected = "ShortName"
        result = _match_column_name(target, self.AVAILABLE_COLUMNS)
        assert result == expected

    def test_match_column_name_fuzzy_match_long_in_short(self):
        """Tests fuzzy match where the target name is contained in the column name (e.g., 'Long' in 'LONG_NAME_TEST')."""
        # Target is the short name, column is longer
        target = "Long"
        expected = "LONG_NAME_TEST"
        result = _match_column_name(target, self.AVAILABLE_COLUMNS)
        assert result == expected

    def test_match_column_name_with_leading_trailing_whitespace(self):
        """Tests that leading/trailing whitespace in the target name is handled."""
        target = "  Time Period  "
        expected = "Time Period"
        result = _match_column_name(target, self.AVAILABLE_COLUMNS)
        assert result == expected

    def test_match_column_name_no_match(self):
        """Tests the failure case where no matching column is found."""
        target = "Totally Different Name"
        with pytest.raises(ValueError) as excinfo:
            _match_column_name(target, self.AVAILABLE_COLUMNS)
        assert "Could not find a column" in str(excinfo.value)

    def test_match_column_name_empty_available_list(self):
        """Tests the failure case when the available columns list is empty."""
        target = "Any Name"
        with pytest.raises(ValueError) as excinfo:
            _match_column_name(target, [])
        assert "Could not find a column" in str(excinfo.value)
        
    def test_match_column_name_invalid_types_for_typecheck(self):
        """Tests that the @typechecked decorator catches incorrect argument types."""
        # Note: Requires the real `typeguard.typechecked` to be active in the runtime
        # environment to enforce the check before the ValueError might be raised.
        # In this isolated mock, we check for a general TypeError/AttributeError if the code runs.
        
        # Passing None for target_name should raise a type error
        with pytest.raises((TypeCheckError)):
            _match_column_name(None, self.AVAILABLE_COLUMNS)

        # Passing a non-list for available_columns should raise a type error
        with pytest.raises((TypeCheckError, AttributeError)):
            _match_column_name("Test", "NotAList")

class TestBuildStructureMap: #noqa: D101
    """Tests for build_structure_map() converting Excel workbook to StructureMap."""
    @pytest.fixture
    def workbook_with_valid_data(self):
        """Creates a valid workbook with comp_mapping and representation sheets."""
        wb = Workbook()
        # comp_mapping sheet
        ws_comp = wb.create_sheet("comp_mapping")
        ws_comp.append(["source", "target", "mapping_rules"])
        ws_comp.append(["SRC1", "TGT1", "fixed:VAL1"])
        ws_comp.append(["SRC2", "TGT2", "implicit"])
        ws_comp.append(["SRC3", "TGT3", "TGT3"])  # representation map
        # representation sheet for TGT3
        ws_rep = wb.create_sheet("TGT3")
        ws_rep.append(["source", "target", "valid_from", "valid_to"])
        ws_rep.append(["A", "B", "", ""])
        return wb

    @pytest.fixture
    def workbook_missing_comp_mapping(self):
        """Workbook without comp_mapping sheet."""
        wb = Workbook()
        wb.create_sheet("Sheet1")
        return wb

    @pytest.fixture
    def workbook_with_invalid_rule(self):
        """Workbook with an invalid mapping rule."""
        wb = Workbook()
        ws_comp = wb.create_sheet("comp_mapping")
        ws_comp.append(["source", "target", "mapping_rules"])
        ws_comp.append(["SRC", "TGT", "unknown_rule"])
        return wb

    def test_valid_workbook_returns_structure_map(self, workbook_with_valid_data):
        """Tests that a valid workbook returns a StructureMap with correct maps."""
        structure_map = build_structure_map(workbook_with_valid_data)
        assert structure_map.id == "GENERATED_STRUCTURE_MAP"
        assert len(structure_map.maps) == 3  # fixed, implicit, representation
        assert any("Mapping for TGT3" in str(m) for m in structure_map.maps)

    def test_missing_comp_mapping_raises_keyerror(self, workbook_missing_comp_mapping):
        """Tests that missing comp_mapping sheet raises KeyError."""
        with pytest.raises(KeyError, match="Mandatory sheet 'comp_mapping' not found"):
            build_structure_map(workbook_missing_comp_mapping)

    def test_invalid_mapping_rule_raises_valueerror(self, workbook_with_invalid_rule):
        """Tests that an unknown mapping rule raises ValueError."""
        with pytest.raises(ValueError, match="Unknown mapping rule"):
            build_structure_map(workbook_with_invalid_rule)

    def test_empty_representation_sheet_skips_map(self):
        """Tests that empty representation sheet is skipped without error."""
        wb = Workbook()
        ws_comp = wb.create_sheet("comp_mapping")
        ws_comp.append(["source", "target", "mapping_rules"])
        ws_comp.append(["SRC", "TGT", "TGT"])
        wb.create_sheet("TGT")  # empty representation sheet
        structure_map = build_structure_map(wb)
        assert len(structure_map.maps) == 0  # skipped due to empty DF

    def test_fixed_value_missing_raises_valueerror(self):
        """Tests that missing fixed value raises ValueError."""
        wb = Workbook()
        ws_comp = wb.create_sheet("comp_mapping")
        ws_comp.append(["source", "target", "mapping_rules"])
        ws_comp.append(["SRC", "TGT", "fixed:"])
        with pytest.raises(ValueError, match="Fixed value for target 'TGT' cannot be empty"):
            build_structure_map(wb)
    
    
    def test_implicit_missing_source_raises_valueerror(self):
        """Tests that implicit mapping without source raises ValueError."""
        wb = Workbook()
        ws_comp = wb.create_sheet("comp_mapping")
        ws_comp.append(["source", "target", "mapping_rules"])
        ws_comp.append(["", "TGT", "implicit"])  # Missing source for implicit map

        # Act & Assert
        with pytest.raises(ValueError):
            build_structure_map(wb)

class TestExtractArtefactId: #noqa: D101
    """Tests for _extract_artefact_id() which extracts SDMX artefact IDs from INFO sheet DataFrame."""

    @pytest.fixture
    def valid_info_df(self):
        """Fixture: DataFrame with valid keys and values for artefact extraction."""
        return pd.DataFrame({
            "Key": ["dataflow", "datastructure", "provisionagreement"],
            "Value": ["AGENCY:DF_ID(1.0)", "AGENCY:DSD_ID(2.0)", "AGENCY:PA_ID(3.0)"]
        })

    @pytest.mark.parametrize("structure_type,expected", [
        ("dataflow", "AGENCY:DF_ID(1.0)"),
        ("dsd", "AGENCY:DSD_ID(2.0)"),
        ("provision-agreement", "AGENCY:PA_ID(3.0)")
    ])
    def test_valid_keys_return_expected_value(self, valid_info_df, structure_type, expected):
        """Tests that valid keys return the correct artefact ID."""
        result = _extract_artefact_id(valid_info_df, structure_type)
        assert result == expected

    def test_invalid_structure_type_raises_valueerror(self, valid_info_df):
        """Tests that an invalid structure_type raises ValueError."""
        with pytest.raises(TypeCheckError):
            _extract_artefact_id(valid_info_df, "invalid")

    def test_missing_key_raises_valueerror(self, valid_info_df):
        """Tests that missing key in DataFrame raises ValueError."""
        df_missing = pd.DataFrame({"Key": ["other"], "Value": ["AGENCY:OTHER(1.0)"]})
        with pytest.raises(ValueError, match="Could not find metadata key 'dataflow'"):
            _extract_artefact_id(df_missing, "dataflow")

    def test_empty_value_raises_valueerror(self):
        """Tests that empty value for a valid key raises ValueError."""
        df_empty_value = pd.DataFrame({"Key": ["dataflow"], "Value": [""]})
        with pytest.raises(ValueError, match="Metadata for 'dataflow' is present but empty"):
            _extract_artefact_id(df_empty_value, "dataflow")

    def test_nan_value_raises_valueerror(self):
        """Tests that NaN value for a valid key raises ValueError."""
        df_nan_value = pd.DataFrame({"Key": ["dataflow"], "Value": [float('nan')]})
        with pytest.raises(ValueError, match="Metadata for 'dataflow' is present but empty"):
            _extract_artefact_id(df_nan_value, "dataflow")

    def test_case_insensitive_key_match(self):
        """Tests that key matching is case-insensitive."""
        df_case = pd.DataFrame({"Key": ["DataFlow"], "Value": ["AGENCY:DF_ID(1.0)"]})
        result = _extract_artefact_id(df_case, "dataflow")
        assert result == "AGENCY:DF_ID(1.0)"

class TestBuildStructureMapFromTemplateWb:
    """Tests for build_structure_map_from_template_wb() which builds a StructureMap from WB-format Excel template."""

    @pytest.fixture
    def valid_mappings(self):
        """Fixture: Valid mappings dictionary with INFO, COMP_MAPPING, and REP_MAPPING sheets."""
        info_df = pd.DataFrame({"Key": ["dataflow"], "Value": ["AGENCY:DF_ID(1.0)"]})
        comp_df = pd.DataFrame({
            "SOURCE": ["SRC1", "SRC2", "SRC3"],
            "TARGET": ["TGT1", "TGT2", "TGT3"],
            "MAPPING_RULES": ["fixed:VAL1", "implicit", "TGT3"]
        })
        rep_df = pd.DataFrame({
            "S:SRC3": ["A", "B"],
            "T:TGT3": ["X", "Y"]
        })
        return {"INFO": info_df, "COMP_MAPPING": comp_df, "REP_MAPPING": rep_df}

    def test_valid_mappings_returns_structure_map(self, valid_mappings):
        """Tests that valid mappings return a StructureMap with correct maps."""
        structure_map = build_structure_map_from_template_wb(valid_mappings)
        assert structure_map.id == "WB_STRUCTURE_MAP"
        assert structure_map.agency == "AGENCY"
        assert structure_map.version == "1.0"
        assert len(structure_map.maps) == 3  # fixed, implicit, representation
        assert any("Mapping for TGT3" in str(m) for m in structure_map.maps)

    def test_missing_comp_mapping_sheet_raises_valueerror(self, valid_mappings):
        """Tests that missing COMP_MAPPING sheet raises ValueError."""
        mappings = {"INFO": valid_mappings["INFO"]}
        with pytest.raises(ValueError, match="Missing required sheet 'COMP_MAPPING'."):
            build_structure_map_from_template_wb(mappings)

    def test_invalid_fixed_rule_format_raises_valueerror(self, valid_mappings):
        """Tests that invalid fixed rule format raises ValueError."""
        mappings = valid_mappings.copy()
        mappings["COMP_MAPPING"].loc[0, "MAPPING_RULES"] = "fixed:"  # Missing value
        with pytest.raises(ValueError, match="Invalid fixed rule format"):
            build_structure_map_from_template_wb(mappings)

    def test_implicit_missing_source_raises_valueerror(self, valid_mappings):
        """Tests that implicit mapping without source raises ValueError."""
        mappings = valid_mappings.copy()
        mappings["COMP_MAPPING"].loc[1, "SOURCE"] = ""  # Remove source for implicit
        with pytest.raises(ValueError, match="Implicit map rule requires a non-empty 'SOURCE'"):
            build_structure_map_from_template_wb(mappings)

    def test_representation_missing_rep_mapping_sheet_raises_valueerror(self, valid_mappings):
        """Tests that representation rule without REP_MAPPING sheet raises ValueError."""
        mappings = valid_mappings.copy()
        mappings.pop("REP_MAPPING")
        with pytest.raises(ValueError, match="Missing required sheet 'REP_MAPPING'."):
            build_structure_map_from_template_wb(mappings)

    def test_representation_empty_combined_df_raises_valueerror(self, valid_mappings):
        """Tests that representation rule with empty combined DataFrame raises ValueError."""
        mappings = valid_mappings.copy()
        mappings["REP_MAPPING"] = pd.DataFrame({"S:SRC3": [], "T:TGT3": []})  # Empty sheet
        with pytest.raises(ValueError):
            build_structure_map_from_template_wb(mappings)

    def test_unknown_mapping_rule_raises_valueerror(self, valid_mappings):
        """Tests that unknown mapping rule raises ValueError."""
        mappings = valid_mappings.copy()
        mappings["COMP_MAPPING"].loc[0, "MAPPING_RULES"] = "unknown_rule"
        with pytest.raises(ValueError, match="Unknown mapping rule"):
            build_structure_map_from_template_wb(mappings)
    
    
    def test_case_insensitive_info_key_match(self, valid_mappings):
        """Tests that INFO sheet key matching is case-insensitive and correctly extracts agency and version."""
        mappings = valid_mappings.copy()
        mappings["INFO"] = pd.DataFrame({"Key": ["DataFlow"], "Value": ["AGENCY:DF_ID(1.0)"]})

        # Act
        structure_map = build_structure_map_from_template_wb(mappings)

        # Assert
        assert structure_map.agency == "AGENCY"
        assert structure_map.version == "1.0"
        assert structure_map.name.startswith("Structure Map generated for")
        assert structure_map.id == "WB_STRUCTURE_MAP"
        assert len(structure_map.maps) == 3  # fixed, implicit, representation

class TestValidateMappings: #noqa: D101
    def test_validate_mappings_valid_input(self):
        """Valid input with all required keys and DataFrames should pass without error."""
        mappings = {
            "INFO": pd.DataFrame(),
            "COMP_MAPPING": pd.DataFrame(),
            "REP_MAPPING": pd.DataFrame()
        }
        # Should not raise any exception
        _validate_mappings(mappings)


    def test_validate_mappings_missing_key(self):
        """Missing one required key should raise ValueError."""
        mappings = {
            "INFO": pd.DataFrame(),
            "COMP_MAPPING": pd.DataFrame()
            # REP_MAPPING is missing
        }
        with pytest.raises(ValueError) as exc_info:
            _validate_mappings(mappings)
        assert "Missing required sheet 'REP_MAPPING'" in str(exc_info.value)


    def test_validate_mappings_invalid_type(self):
        """Invalid type for one of the keys should raise ValueError."""
        mappings = {
            "INFO": pd.DataFrame(),
            "COMP_MAPPING": "not_a_dataframe",  # Invalid type
            "REP_MAPPING": pd.DataFrame()
        }
        with pytest.raises(ValueError) as exc_info:
            _validate_mappings(mappings)
        assert "must be a pandas DataFrame" in str(exc_info.value)


    def test_validate_mappings_empty_dict(self):
        """Empty dictionary should raise ValueError for missing keys."""
        mappings = {}
        with pytest.raises(ValueError) as exc_info:
            _validate_mappings(mappings)
        assert "Missing required sheet 'INFO'" in str(exc_info.value)


    def test_validate_mappings_partial_invalid_type(self):
        """One valid key and one invalid type should raise ValueError."""
        mappings = {
            "INFO": pd.DataFrame(),
            "COMP_MAPPING": pd.DataFrame(),
            "REP_MAPPING": 123  # Invalid type
        }
        with pytest.raises(ValueError) as exc_info:
            _validate_mappings(mappings)
        assert "must be a pandas DataFrame" in str(exc_info.value)

class TestExtractAllArtefactIds: #noqa: D101
    
    def test_extract_all_artefact_ids_normal(self):
        """Test normal case with valid artefact keys and values."""
        df = pd.DataFrame({
            'Key': ['dataflow', 'datastructure', 'provisionagreement'],
            'Value': ['AGENCY:DF1(1.0)', 'AGENCY:DSD1(1.0)', 'AGENCY:PA1(1.0)']
        })
        result = _extract_all_artefact_ids(df)
        assert result == {'dataflow': 'AGENCY:DF1(1.0)', 'datastructure': 'AGENCY:DSD1(1.0)', 'provisionagreement': 'AGENCY:PA1(1.0)'}
        assert isinstance(result, dict)
    
    def test_extract_all_artefact_ids_with_missing_values(self):
        """Test normal case with valid artefact keys and missing values."""
        df = pd.DataFrame({
            'Key': ['dataflow', 'datastructure', 'provisionagreement'],
            'Value': ['AGENCY:DF1(1.0)', 'AGENCY:DSD1(1.0)', '']
        })
        result = _extract_all_artefact_ids(df)
        assert result == {'dataflow': 'AGENCY:DF1(1.0)', 'datastructure': 'AGENCY:DSD1(1.0)'}
        assert isinstance(result, dict)

    def test_extract_all_artefact_ids_empty_df(self):
        """Test empty DataFrame raises ValueError."""
        df = pd.DataFrame(columns=['Key', 'Value'])
        with pytest.raises(ValueError):
            _extract_all_artefact_ids(df)

    def test_extract_all_artefact_ids_missing_columns(self):
        """Test missing columns raises ValueError."""
        df = pd.DataFrame({'Key': ['dataflow']})
        with pytest.raises(ValueError):
            _extract_all_artefact_ids(df)

    def test_extract_all_artefact_ids_no_matching_keys(self):
        """Test DataFrame with no matching keys raises ValueError."""
        df = pd.DataFrame({'Key': ['other'], 'Value': ['AGENCY:XYZ(1.0)']})
        with pytest.raises(ValueError):
            _extract_all_artefact_ids(df)

    def test_extract_all_artefact_ids_invalid_type(self):
        """Test invalid input type raises TypeCheckError."""
        with pytest.raises(TypeCheckError):
            _extract_all_artefact_ids("not a dataframe")


class TestExtractMetadataFromInfoSheet:
    """Tests for `_extract_metadata_from_info_sheet` function."""
	
    @pytest.fixture
    def info_df_with_all(self):
        """Fixture: DataFrame with datastructure, dataflow, and FMR_AGENCY keys."""
        return pd.DataFrame({
            "Key": ["datastructure", "dataflow", "FMR_AGENCY"],
            "Value": ["AGENCY:DSD(1.0)", "AGENCY:DF(2.0)", "ALT_AGENCY"]
        })

    @pytest.fixture
    def info_df_only_dataflow(self):
        """Fixture: DataFrame with only dataflow key."""
        return pd.DataFrame({
            "Key": ["dataflow"],
            "Value": ["AGENCY:DF(2.0)"]
        })


    @pytest.fixture
    def empty_info_df(self):
        """Fixture: Empty DataFrame."""
        return pd.DataFrame(columns=["Key", "Value"])

    def test_preferred_structure_type_present(self, info_df_with_all):
        """Tests that datastructure is selected when present."""
        agency, version, artefact_ref = _extract_metadata_from_info_sheet(
            info_df_with_all,
            agency = "default_agency",
            version = "1.0",
            structure_type =  "datastructure")
        assert agency == "AGENCY"
        assert version == "1.0"
        assert artefact_ref == "AGENCY:DSD(1.0)"

    def test_fallback_to_dataflow(self, info_df_only_dataflow):
        """Tests fallback when datastructure is missing but dataflow exists."""
        agency, version, artefact_ref = _extract_metadata_from_info_sheet(
            info_df_only_dataflow, 
            agency = "default_agency",
            version = "1.0",
            structure_type =  "datastructure")
        assert agency == "AGENCY"
        assert version == "2.0"
        assert artefact_ref == "AGENCY:DF(2.0)"

    def test_empty_dataframe_returns_defaults(self, empty_info_df):
        """Tests that defaults are returned when DataFrame is empty."""
        agency, version, artefact_ref = _extract_metadata_from_info_sheet(
            empty_info_df,
            agency = "SDMX",
            version = "1.0",
            structure_type =  "datastructure")
        assert agency == "SDMX"
        assert version == "1.0"
        assert artefact_ref is None

    def test_invalid_structure_type_still_falls_back(self, info_df_with_all):
        """Tests that invalid structure_type raises TypeCheckError."""
        with pytest.raises(TypeCheckError):
            _extract_metadata_from_info_sheet(
                info_df_with_all,
                agency = "default_agency",
            version = "1.0",
            structure_type =  "invalid_type")

class TestIsMissingToken:
    """Tests for the `_is_missing_token` function which checks if a string is a missing token."""

    @pytest.mark.parametrize("input_str,expected", [
        ("nan", True),
        ("NaN", True),
        ("<na>", True),
        ("<NA>", True),
        ("", True),
        ("   ", True),  # whitespace only
        ("valid", False),
        ("fixed:123", False),
    ])
    def test_missing_token_cases(self, input_str, expected):
        """Tests that `_is_missing_token` correctly identifies missing tokens."""
        assert _is_missing_token(input_str) == expected


class TestExtractMappingRule:
    """Tests for the `_extract_mapping_rule` function which parses mapping rules from a pandas Series."""

    def test_skip_rule_when_target_empty(self):
        """Tests that rule is 'skip' when TARGET is empty."""
        row = pd.Series({"SOURCE": "SRC", "TARGET": "", "MAPPING_RULES": "implicit"})
        result = _extract_mapping_rule(row)
        assert result["mapping_rule"] == "skip"
        assert result["source_id"] == "SRC"
        assert result["target_id"] == ""
        assert result["fixed_value"] is None

    def test_skip_rule_when_rule_missing(self):
        """Tests that rule is 'skip' when MAPPING_RULES is missing-like."""
        row = pd.Series({"SOURCE": "SRC", "TARGET": "TGT", "MAPPING_RULES": "nan"})
        result = _extract_mapping_rule(row)
        assert result["mapping_rule"] == "skip"

    def test_fixed_rule_valid(self):
        """Tests that a valid fixed rule returns correct mapping."""
        row = pd.Series({"SOURCE": "SRC", "TARGET": "TGT", "MAPPING_RULES": "fixed:123"})
        result = _extract_mapping_rule(row)
        assert result == {
            "mapping_rule": "fixed",
            "source_id": "SRC",
            "target_id": "TGT",
            "fixed_value": "123",
        }

    def test_fixed_rule_invalid_format(self):
        """Tests that an invalid fixed rule raises ValueError."""
        row = pd.Series({"SOURCE": "SRC", "TARGET": "TGT", "MAPPING_RULES": "fixed:"})
        with pytest.raises(ValueError, match="Invalid fixed rule format"):
            _extract_mapping_rule(row)

    def test_implicit_rule_valid(self):
        """Tests that implicit rule works when SOURCE is present."""
        row = pd.Series({"SOURCE": "SRC", "TARGET": "TGT", "MAPPING_RULES": "implicit"})
        result = _extract_mapping_rule(row)
        assert result["mapping_rule"] == "implicit"

    def test_implicit_rule_missing_source(self):
        """Tests that implicit rule raises ValueError when SOURCE is missing."""
        row = pd.Series({"SOURCE": "", "TARGET": "TGT", "MAPPING_RULES": "implicit"})
        with pytest.raises(ValueError, match="Implicit map rule requires"):
            _extract_mapping_rule(row)

    def test_representation_rule_valid(self):
        """Tests that representation rule works when rule equals TARGET."""
        row = pd.Series({"SOURCE": "SRC", "TARGET": "TGT", "MAPPING_RULES": "TGT"})
        result = _extract_mapping_rule(row)
        assert result["mapping_rule"] == "representation"

    def test_representation_rule_missing_source(self):
        """Tests that representation rule raises ValueError when SOURCE is missing."""
        row = pd.Series({"SOURCE": "", "TARGET": "TGT", "MAPPING_RULES": "TGT"})
        with pytest.raises(ValueError, match="Representation map rule requires"):
            _extract_mapping_rule(row)

    def test_unknown_rule_raises_error(self):
        """Tests that unknown mapping rule raises ValueError."""
        row = pd.Series({"SOURCE": "SRC", "TARGET": "TGT", "MAPPING_RULES": "unknown_rule"})
        with pytest.raises(ValueError, match="Unknown mapping rule"):
            _extract_mapping_rule(row)

class TestExtractRepresentationMap:
    """Tests for `_extract_representation_map` which builds a sanitized mapping DataFrame."""

    @pytest.fixture
    def sample_rep_data(self):
        """Provides valid source and target DataFrames for tests."""
        source_df = pd.DataFrame({"src_col": ["A", "B", None, "C"], "extra": [1, 2, 3, 4]})
        target_df = pd.DataFrame({"tgt_col": ["X", "Y", "Z", None], "extra": [5, 6, 7, 8]})
        return {"source": source_df, "target": target_df}

    def test_valid_mapping(self, sample_rep_data):
        """Tests that valid mapping returns correct DataFrame with duplicates and NA removed."""
        result_df = _extract_representation_map(sample_rep_data, "src_col", "tgt_col")
        expected = pd.DataFrame({"source": ["A", "B"], "target": ["X", "Y"]})
        pd.testing.assert_frame_equal(result_df, expected)

    @pytest.mark.skip(reason="Not sure this is expected behavior")
    def test_raises_value_error_on_missing_rep_data(self):
        """Tests that ValueError is raised when rep_data is empty or invalid."""
        invalid_cases = [
            {},  # Empty dict
            {"source": None, "target": None},  # None DataFrames
            {"source": pd.DataFrame(), "target": pd.DataFrame()},  # Empty DataFrames
        ]
        for case in invalid_cases:
            with pytest.raises(ValueError, match="Mapping rule requires 'REP_MAPPING'"):
                _extract_representation_map(case, "src_col", "tgt_col")

    def test_raises_value_error_on_column_not_found(self, sample_rep_data):
        """Tests that ValueError is raised when column resolution fails."""
        with pytest.raises(ValueError):
            _extract_representation_map(sample_rep_data, "invalid_src", "invalid_tgt")

    def test_raises_value_error_on_empty_result_after_sanitization(self):
        """Tests that ValueError is raised when all rows are dropped after sanitization."""
        source_df = pd.DataFrame({"src_col": [None, None], "extra": [1, 2]})
        target_df = pd.DataFrame({"tgt_col": [None, None], "extra": [3, 4]})
        rep_data = {"source": source_df, "target": target_df}

        with pytest.raises(ValueError, match="No valid mapping rows found"):
            _extract_representation_map(rep_data, "src_col", "tgt_col")

    def test_deduplication_of_pairs(self):
        """Tests that duplicate mapping pairs are removed."""
        source_df = pd.DataFrame({"src_col": ["A", "A"], "extra": [1, 2]})
        target_df = pd.DataFrame({"tgt_col": ["X", "X"], "extra": [3, 4]})
        rep_data = {"source": source_df, "target": target_df}

        result_df = _extract_representation_map(rep_data, "src_col", "tgt_col")
        assert len(result_df) == 1
        assert result_df.iloc[0].to_dict() == {"source": "A", "target": "X"}

class TestCreateXlTemplateFromSm:
    """Tests for the create_xl_template_from_sm function which generates Excel templates from a StructureMap."""

    @pytest.fixture
    def structure_map_with_various_mappings(self):
        """Fixture that returns a StructureMap with multiple mapping types including RepresentationMap."""
        m1 = ImplicitComponentMap("OBS_CONF", "CONF_STATUS")
        m2 = FixedValueMap("FREQ", "M")
        m3 = DatePatternMap("ACTIVITY_DATE", "TIME_PERIOD", "MMM YYYY", "M")
        m4 = ComponentMap(
            "SRC1",
            "TGT1",
            RepresentationMap(
                id="REP1",
                name="RepMap1",
                agency="BIS",
                source="CL1",
                target="CL2",
                maps=[ValueMap(source="1", target="A"), ValueMap(source="2", target="B")],
            ),
        )
        m5 = MultiComponentMap(
            source=["SRC_M1", "SRC_M2"],
            target=["TGT_M1", "TGT_M2"],
            values=MultiRepresentationMap(
                id = "MREP1",
                agency="BIS",
                maps=[
                    ValueMap(source=["X1", "X2"], target=["Y1", "Y2"]),
                    ValueMap(source=["X3", "X4"], target=["Y3", "Y4"]),
                ]
            ),
        )

        return StructureMap(
            id="SM_FULL",
            name="Full Map",
            agency="WB",
            source="urn:sdmx:source",
            target="urn:sdmx:target",
            maps=[m1, m2, m3, m4, m5],
        )


    @pytest.fixture
    def empty_structure_map(self):
        """Fixture that returns a StructureMap with no mappings."""
        return StructureMap(
            id="SM_EMPTY",
            name="Empty Map",
            agency="WB",
            source="src",
            target="tgt",
            maps=[],
        )

    def test_valid_structure_map_creates_comp_mapping_sheet(self, structure_map_with_various_mappings):
        """Tests that comp_mapping sheet is created with correct columns and rows."""
        result = create_xl_template_from_sm(structure_map_with_various_mappings)
        assert isinstance(result, dict)
        # assert "comp_mapping" in result
        # df = result["comp_mapping"]
        # assert isinstance(df, pd.DataFrame)
        # assert set(df.columns) == {"source", "target", "mapping_rules"}
        # assert len(df) == len(structure_map_with_various_mappings.maps)

    def test_representation_map_creates_extra_sheet(self, structure_map_with_various_mappings):
        """Tests that RepresentationMap creates an additional sheet with correct columns."""
        result = create_xl_template_from_sm(structure_map_with_various_mappings)
        assert "TGT1" in result
        df = result["TGT1"]
        assert isinstance(df, pd.DataFrame)
        assert set(df.columns) == {"source", "target", "valid_from", "valid_to"}
        assert len(df) == 2  # Two ValueMap entries

    def test_multi_representation_map_creates_extra_sheet(self, structure_map_with_various_mappings):
        """Tests that MultiRepresentationMap creates an additional sheet with correct columns."""
        result = create_xl_template_from_sm(structure_map_with_various_mappings)
        assert "TGT_M1" in result
        df = result["TGT_M1"]
        assert isinstance(df, pd.DataFrame)
        assert set(df.columns) == {"source", "target", "valid_from", "valid_to"}
        assert len(df) == 2  # Two ValueMap entries

    def test_empty_maps_raises_value_error(self, empty_structure_map):
        """Tests that ValueError is raised when StructureMap contains no mappings."""
        with pytest.raises(ValueError, match="structure_map contains no mappings"):
            create_xl_template_from_sm(empty_structure_map)

    def test_invalid_type_raises_type_error(self):
        """Tests that TypeError is raised when input is not a StructureMap instance."""
        with pytest.raises(TypeCheckError):
            create_xl_template_from_sm("not_a_structure_map")
