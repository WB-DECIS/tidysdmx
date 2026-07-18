"""Tests for tidysdmx.utils module."""

import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from pysdmx.model import (
    Code,
    Codelist,
    Component,
    Components,
    Role,
    Schema,
)
from typeguard import TypeCheckError

from tidysdmx.utils import (
    build_excel_workbook,
    create_mapping_rules,
    extract_component_ids,
    extract_validation_info,
    get_codelist_ids,
    parse_mapping_template_wb,
    sdmx_reference_cols_for,
    write_excel_mapping_template,
)

# region Fixtures


@pytest.fixture
def mapping_template_path() -> Path:
    """Return path to the Excel mapping template fixture."""
    file_path = Path(__file__).parent / "fixtures" / "data" / "wb_mapping_template.xlsx"
    if not file_path.exists():
        raise FileNotFoundError(f"Expected fixture file not found: {file_path}")
    return file_path


@pytest.fixture
def invalid_mapping_template_path() -> Path:
    """Return path to the invalid mapping template fixture."""
    file_path = (
        Path(__file__).parent / "fixtures" / "data" / "wb_mapping_template_invalid.txt"
    )
    if not file_path.exists():
        raise FileNotFoundError(f"Expected fixture file not found: {file_path}")
    return file_path


@pytest.fixture
def sample_codelist() -> Codelist:
    """Create a pysdmx Codelist with two codes."""
    return Codelist(
        id="CL_TEST",
        agency="AGENCY",
        version="1.0",
        name={"en": "Test Codelist"},
        codes=[
            Code(id="CODE_A", name={"en": "Code A"}),
            Code(id="CODE_B", name={"en": "Code B"}),
        ],
    )


@pytest.fixture
def sample_components(sample_codelist: Codelist) -> Components:
    """Create a Components collection with dimension, attribute, and measure."""
    comp_dim = Component(
        id="DIM_MANDATORY",
        required=True,
        role=Role.DIMENSION,
    )
    comp_coded = Component(
        id="CODED",
        required=False,
        role=Role.ATTRIBUTE,
        local_codes=sample_codelist,
    )
    comp_simple = Component(
        id="SIMPLE",
        required=False,
        role=Role.MEASURE,
    )
    return Components([comp_dim, comp_coded, comp_simple])


@pytest.fixture
def sample_schema(sample_components: Components) -> Schema:
    """Create a Schema populated with sample components."""
    return Schema(
        id_="TEST_SCHEMA",
        agency="AGENCY",
        version="1.0",
        name={"en": "Test Schema"},
        context="dataflow",
        components=sample_components,
    )


@pytest.fixture
def test_workbook_data() -> tuple[list[str], list[str]]:
    """Common test data for components and rep_maps."""
    components = ["C_FREQ", "C_REF_AREA", "C_UNIT", "C_OBS_VALUE"]
    # Duplicate to test deduplication
    rep_maps = ["C_REF_AREA", "C_UNIT", "C_REF_AREA"]
    return components, rep_maps


# endregion


class TestExtractComponentIds:
    def test_extract_component_ids_empty(self):
        """Raise ValueError if Schema has no components."""
        empty_schema = Schema(
            id="EMPTY",
            agency="TEST",
            version="1.0",
            context="dataflow",
            components=Components([]),
        )

        with pytest.raises(ValueError, match="Schema contains no components"):
            extract_component_ids(empty_schema)

    def test_extract_component_ids_invalid_type(self):
        """Raise TypeCheckError if input is not a Schema."""
        with pytest.raises(TypeCheckError):
            extract_component_ids("Not a Schema Object")


class TestExtractValidationInfo:
    @pytest.mark.parametrize(
        "invalid_input",
        [None, {}, [], "not_a_schema", 123],
    )
    def test_extract_validation_info(self, invalid_input):
        """Raise TypeCheckError for non-schema input."""
        with pytest.raises(TypeCheckError):
            extract_validation_info(invalid_input)

    @pytest.mark.integration
    def test_extract_validation_has_expected_structure(self, ifpri_asti_schema):
        """Ensure the returned dict has the expected keys and types."""
        result = extract_validation_info(ifpri_asti_schema)

        assert isinstance(result, dict)
        expected_keys = {
            "valid_comp",
            "mandatory_comp",
            "coded_comp",
            "codelist_ids",
            "dim_comp",
            "sdmx_cols",
        }
        assert set(result.keys()) == expected_keys
        assert all(
            isinstance(item, list)
            for key, item in result.items()
            if key != "codelist_ids"
        )
        assert isinstance(result["codelist_ids"], dict)

    def test_extract_validation_has_expected_structure2(self, sdmx_schema):
        """Ensure the returned dict has the expected keys and types."""
        result = extract_validation_info(sdmx_schema)

        assert isinstance(result, dict)
        expected_keys = {
            "valid_comp",
            "mandatory_comp",
            "coded_comp",
            "codelist_ids",
            "dim_comp",
            "sdmx_cols",
        }
        assert set(result.keys()) == expected_keys
        assert all(
            isinstance(item, list)
            for key, item in result.items()
            if key != "codelist_ids"
        )
        assert isinstance(result["codelist_ids"], dict)

    def test_sdmx_cols_inferred_from_dataflow_context(self, sdmx_schema):
        """Dataflow-context schema yields the standard reference columns."""
        result = extract_validation_info(sdmx_schema)
        assert result["sdmx_cols"] == ["STRUCTURE", "STRUCTURE_ID", "ACTION"]

    @pytest.mark.integration
    def test_sdmx_cols_inferred_from_datastructure_context(self, ifpri_asti_schema):
        """Datastructure-context schema yields STRUCTURE reference columns."""
        result = extract_validation_info(ifpri_asti_schema)
        assert result["sdmx_cols"] == ["STRUCTURE", "STRUCTURE_ID", "ACTION"]


class TestSdmxReferenceColsFor:
    @pytest.mark.parametrize(
        "context, expected",
        [
            ("dataflow", ["STRUCTURE", "STRUCTURE_ID", "ACTION"]),
            ("datastructure", ["STRUCTURE", "STRUCTURE_ID", "ACTION"]),
            (
                "provisionagreement",
                ["STRUCTURE", "STRUCTURE_ID", "ACTION"],
            ),
        ],
    )
    def test_returns_expected_columns(self, context, expected):
        """Every SDMX context maps to the standard SDMX-CSV reference columns."""
        assert sdmx_reference_cols_for(context) == expected

    def test_raises_on_invalid_context(self):
        """Unknown contexts raise a TypeCheckError (rejected by typeguard)."""
        with pytest.raises(TypeCheckError):
            sdmx_reference_cols_for("not_a_context")


class TestGetCodelistIds:
    @pytest.mark.integration
    def test_get_codelist_ids_has_expected_structure(self, ifpri_asti_schema):
        """Ensure returned dict maps coded components to code ID lists."""
        comp = ifpri_asti_schema.components
        coded_comp = [c.id for c in comp if comp[c.id].local_codes is not None]

        result = get_codelist_ids(comp, coded_comp)
        assert isinstance(result, dict)
        for key, value in result.items():
            assert key in coded_comp
            assert isinstance(value, list)
            assert all(isinstance(code_id, str) for code_id in value)

    @pytest.mark.skip(reason="Test needs to be modified to use correct inputs")
    def test_get_codelist_ids(self):
        """Test get_codelist_ids with simple inputs."""
        comp = {"dim1": "Dimension 1", "dim2": "Dimension 2"}
        coded_comp = {"dim1": ["A", "B"], "dim2": ["C", "D"]}
        expected_output = {"dim1": ["A", "B"], "dim2": ["C", "D"]}
        assert get_codelist_ids(comp, coded_comp) == expected_output


class TestExtractCodelistIds:
    @pytest.mark.skip(reason="Temporary skipping to generate a coverage report")
    def test_extract_component_ids_normal(self):
        """Retrieve IDs from a valid schema with multiple components."""
        comp1 = Component(id="FREQ")
        comp2 = Component(id="TIME_PERIOD")
        schema = Schema(
            context="datastructure",
            agency="ECB",
            id_="EXR",
            components=Components([comp1, comp2]),
            version="1.0.0",
            urns=[],
        )
        result = extract_component_ids(schema)
        assert result == ["FREQ", "TIME_PERIOD"]
        assert all(isinstance(cid, str) for cid in result)

    @pytest.mark.skip(reason="Temporary skipping to generate a coverage report")
    def test_extract_component_ids_single_component(self):
        """Schema with a single component returns a one-element list."""
        comp = Component(id="OBS_VALUE")
        schema = Schema(
            context="datastructure",
            agency="ECB",
            id_="EXR",
            components=Components([comp]),
            version="1.0.0",
            urns=[],
        )
        result = extract_component_ids(schema)
        assert result == ["OBS_VALUE"]
        assert len(result) == 1

    @pytest.mark.skip(reason="Temporary skipping to generate a coverage report")
    def test_extract_component_ids_empty(self):
        """Schema with no components raises ValueError."""
        schema = Schema(
            context="datastructure",
            agency="ECB",
            id_="EXR",
            components=Components([]),
            version="1.0.0",
            urns=[],
        )
        with pytest.raises(ValueError):
            extract_component_ids(schema)

    def test_extract_component_ids_invalid_type(self):
        """Non-Schema input raises TypeCheckError."""
        with pytest.raises(TypeCheckError):
            extract_component_ids("not_a_schema")

    @pytest.mark.skip(reason="Temporary skipping to generate a coverage report")
    def test_extract_component_ids_component_without_id(self):
        """Component without an ID should raise Error."""
        comp = Component(id=None)  # Simulate missing ID
        schema = Schema(
            context="datastructure",
            agency="ECB",
            id_="EXR",
            components=Components([comp]),
            version="1.0.0",
            urns=[],
        )
        with pytest.raises(TypeError):
            extract_component_ids(schema)


class TestCreateMappingRules:
    def test_create_mapping_rules_normal_case(self):
        """Test a mix of matching and non-matching components."""
        components = ["D1", "D2", "D3", "D4"]
        rep_maps = {"D2", "D4"}
        expected = [
            "",
            '=HYPERLINK("#D2!A1","D2")',
            "",
            '=HYPERLINK("#D4!A1","D4")',
        ]
        result = create_mapping_rules(components, rep_maps)
        assert result == expected
        assert isinstance(result, list)
        assert all(isinstance(r, str) for r in result)

    def test_create_mapping_rules_no_matches(self):
        """No component present in rep_maps yields all empty strings."""
        components = ["D1", "D2", "D3"]
        rep_maps = {"D4", "D5"}
        expected = ["", "", ""]
        result = create_mapping_rules(components, rep_maps)
        assert result == expected

    def test_create_mapping_rules_all_matches(self):
        """All components in rep_maps yields all hyperlinks."""
        components = ["D1", "D2", "D3"]
        rep_maps = {"D1", "D2", "D3", "D4"}
        expected = [
            '=HYPERLINK("#D1!A1","D1")',
            '=HYPERLINK("#D2!A1","D2")',
            '=HYPERLINK("#D3!A1","D3")',
        ]
        result = create_mapping_rules(components, rep_maps)
        assert result == expected

    def test_create_mapping_rules_empty_components(self):
        """Empty components list returns empty list."""
        components: list[str] = []
        rep_maps = {"D1", "D2"}
        result = create_mapping_rules(components, rep_maps)
        assert result == []

    def test_create_mapping_rules_none_rep_maps(self):
        """None rep_maps yields all empty strings."""
        components = ["D1", "D2", "D3"]
        result = create_mapping_rules(components, None)
        assert result == ["", "", ""]

    def test_create_mapping_rules_empty_rep_maps(self):
        """Empty set rep_maps yields all empty strings."""
        components = ["D1", "D2", "D3"]
        rep_maps: set[str] = set()
        result = create_mapping_rules(components, rep_maps)
        assert result == ["", "", ""]

    def test_create_mapping_rules_type_error_for_components(self):
        """TypeCheckError when components is not a Sequence[str]."""
        with pytest.raises(TypeCheckError):
            create_mapping_rules(123, {"D2"})  # type: ignore

    def test_create_mapping_rules_type_error_for_rep_maps(self):
        """TypeCheckError when rep_maps is not a Set[str] or None."""
        components = ["D1", "D2"]
        with pytest.raises(TypeCheckError):
            create_mapping_rules(components, [1, 2])  # type: ignore


class TestBuildExcelWorkbook:
    def test_build_excel_workbook_content_and_sheets(self, test_workbook_data):
        """Test successful workbook creation and content structure."""
        components, rep_maps = test_workbook_data

        wb = build_excel_workbook(components, rep_maps)

        # Check sheet names (2 unique rep_maps + 1 default sheet)
        expected_sheet_titles = {"comp_mapping", "C_REF_AREA", "C_UNIT"}
        assert set(wb.sheetnames) == expected_sheet_titles

        # Check default sheet content (comp_mapping)
        main_sheet = wb["comp_mapping"]
        header = [cell.value for cell in main_sheet[1]]
        assert header == ["source", "target", "mapping_rules"]

        # Check mapping_rules column (Column C, Rows 2-5)
        rules_cells = [main_sheet[f"C{i}"].value for i in range(2, 6)]
        expected_rules = [
            "",
            '=HYPERLINK("#C_REF_AREA!A1","C_REF_AREA")',
            '=HYPERLINK("#C_UNIT!A1","C_UNIT")',
            "",
        ]
        assert rules_cells == expected_rules

        # Check rep_map sheet headers
        rep_sheet = wb["C_REF_AREA"]
        rep_header = [cell.value for cell in rep_sheet[1]]
        assert rep_header == [
            "source",
            "target",
            "valid_from",
            "valid_to",
        ]

    def test_build_excel_workbook_no_rep_maps(self, test_workbook_data):
        """Workbook with None rep_maps has only the default sheet."""
        components, _ = test_workbook_data
        wb = build_excel_workbook(components, None)

        assert wb.sheetnames == ["comp_mapping"]

        main_sheet = wb["comp_mapping"]
        rules_cells = [main_sheet[f"C{i}"].value for i in range(2, 6)]
        assert rules_cells == ["", "", "", ""]


class TestWriteExcelMappingTemplate:
    def test_write_excel_mapping_template_success(self, test_workbook_data, tmp_path):
        """Test successful file creation."""
        components, rep_maps = test_workbook_data
        output_path = tmp_path / "test_saved_file.xlsx"

        result_path = write_excel_mapping_template(components, rep_maps, output_path)

        assert result_path == output_path
        assert output_path.exists()
        assert output_path.name == "test_saved_file.xlsx"
        assert os.path.getsize(output_path) > 100

    def test_write_non_existent_dir_raises(self, test_workbook_data, tmp_path):
        """Non-existent parent directory raises FileNotFoundError."""
        components, rep_maps = test_workbook_data
        non_existent_dir = tmp_path / "sub_dir"
        output_path = non_existent_dir / "test_missing_dir.xlsx"

        assert not non_existent_dir.exists()

        with pytest.raises(FileNotFoundError) as excinfo:
            write_excel_mapping_template(components, rep_maps, output_path)

        assert "does not exist" in str(excinfo.value)
        assert not output_path.exists()

    def test_write_excel_mapping_template_integrity_check(
        self, test_workbook_data, tmp_path
    ):
        """Verify the saved file content is correct."""
        components, rep_maps = test_workbook_data
        output_path = tmp_path / "test_integrity.xlsx"

        write_excel_mapping_template(components, rep_maps, output_path)

        wb = load_workbook(output_path)
        expected_sheet_titles = {
            "comp_mapping",
            "C_REF_AREA",
            "C_UNIT",
        }
        assert set(wb.sheetnames) == expected_sheet_titles

        main_sheet = wb["comp_mapping"]
        cell_value = main_sheet["C3"].value
        assert cell_value == ('=HYPERLINK("#C_REF_AREA!A1","C_REF_AREA")')


class TestParseMappingTemplateWb:
    def test_parse_mapping_template_wb_valid(self, mapping_template_path):
        """Test parsing a valid mapping template."""
        result = parse_mapping_template_wb(mapping_template_path)
        assert isinstance(result, dict)
        assert "REP_MAPPING" in result and "COMP_MAPPING" in result
        assert isinstance(result["REP_MAPPING"], pd.DataFrame)
        assert isinstance(result["COMP_MAPPING"], pd.DataFrame)

    def test_parse_mapping_template_wb_file_not_found(self):
        """FileNotFoundError is raised for non-existent file."""
        with pytest.raises(FileNotFoundError):
            parse_mapping_template_wb("non_existent.xlsx")

    def test_parse_mapping_template_wb_invalid_file_type(
        self, invalid_mapping_template_path
    ):
        """ValueError is raised for invalid file type."""
        with pytest.raises(ValueError):
            parse_mapping_template_wb(invalid_mapping_template_path)
